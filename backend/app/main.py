import time
import hashlib
import uuid
import csv
import io
import zipfile
import re
import httpx
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session, selectinload
from sqlalchemy.exc import IntegrityError
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy import text
from sqlalchemy import select, func, and_, or_, case, delete
from datetime import datetime, timedelta, date

from .db import Base, engine, get_db
from . import models, schemas, crud
from .fifo import fifo_allocate

app = FastAPI(title="Print ERP MVP")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/health")
def health():
    return {"status": "ok", "utc": datetime.utcnow().isoformat()}

@app.on_event("startup")
def startup():
    # wait for db (max ~60s)
    for _ in range(60):
        try:
            with engine.connect() as conn:
                conn.execute(text("SELECT 1"))
            break
        except Exception:
            time.sleep(1)
    Base.metadata.create_all(bind=engine)

    # Postgres-only safety rails: anti-duplicate + immutability for money facts
    if engine.dialect.name == "postgresql":
        with engine.connect() as conn:
            # widen ymarket_orders.order_id to BIGINT (YM order ids can be > 2^31)
            try:
                dtype = conn.execute(text("""
                    SELECT data_type FROM information_schema.columns
                    WHERE table_schema = 'public' AND table_name = 'ymarket_orders' AND column_name = 'order_id'
                """)).scalar()
                if dtype == "integer":
                    conn.execute(text("ALTER TABLE ymarket_orders ALTER COLUMN order_id TYPE bigint USING order_id::bigint;"))
            except Exception:
                pass

            # strong anti-dub for imports without external_id
            conn.execute(text(
                "CREATE UNIQUE INDEX IF NOT EXISTS uq_moneyop_fingerprint_notnull "
                "ON money_operations(hash_fingerprint) WHERE hash_fingerprint IS NOT NULL;"
            ))
            # prevent updates/deletes of immutable money facts (allow only void flags)
            conn.execute(text("""
CREATE OR REPLACE FUNCTION prevent_moneyop_mutation()
RETURNS trigger AS $$
BEGIN
  IF TG_OP = 'DELETE' THEN
    RAISE EXCEPTION 'MoneyOperation facts are immutable: delete is запрещен';
  END IF;

  IF (OLD.account_id IS DISTINCT FROM NEW.account_id)
     OR (OLD.transfer_group_id IS DISTINCT FROM NEW.transfer_group_id)
     OR (OLD.posted_at IS DISTINCT FROM NEW.posted_at)
     OR (OLD.amount IS DISTINCT FROM NEW.amount)
     OR (OLD.currency IS DISTINCT FROM NEW.currency)
     OR (OLD.counterparty IS DISTINCT FROM NEW.counterparty)
     OR (OLD.description IS DISTINCT FROM NEW.description)
     OR (OLD.operation_type IS DISTINCT FROM NEW.operation_type)
     OR (OLD.external_id IS DISTINCT FROM NEW.external_id)
     OR (OLD.source IS DISTINCT FROM NEW.source)
     OR (OLD.raw_payload IS DISTINCT FROM NEW.raw_payload)
     OR (OLD.hash_fingerprint IS DISTINCT FROM NEW.hash_fingerprint)
  THEN
    RAISE EXCEPTION 'MoneyOperation facts are immutable: update запрещен';
  END IF;

  RETURN NEW;
END;
$$ LANGUAGE plpgsql;
"""))
            conn.execute(text("""
DO $$
BEGIN
  IF NOT EXISTS (
    SELECT 1 FROM pg_trigger WHERE tgname = 'trg_moneyop_immutable'
  ) THEN
    CREATE TRIGGER trg_moneyop_immutable
    BEFORE UPDATE OR DELETE ON money_operations
    FOR EACH ROW EXECUTE FUNCTION prevent_moneyop_mutation();
  END IF;
END $$;
"""))
            conn.commit()

    # Ensure базовые счета/категории есть всегда (идемпотентно)
    # (помогает, если остался старый docker volume и пользователь не делал down -v)
    with Session(engine) as db:
        ensure_money_bootstrap(db)


def ensure_money_bootstrap(db: Session):
    """Idempotent bootstrap for Money ledger essentials.

    Важно: у пользователя может остаться старая база/volume.
    Мы делаем так, чтобы в интерфейсе всегда были счета/категории,
    даже если он не делал `docker compose down -v`.
    """
    try:
        # Accounts: upsert by name (safe). We don't enforce uniqueness by type.
        seed_accounts = [
            ("bank", "Банк (р/с)"),
            ("cash", "Касса (наличка)"),
            ("marketplace", "WB баланс"),
            ("marketplace", "Ozon баланс"),
            ("marketplace", "Яндекс Маркет баланс"),
            ("acquiring", "Эквайринг (онлайн-оплата)"),
        ]
        for t, n in seed_accounts:
            ex = db.execute(select(models.MoneyAccount).where(func.lower(models.MoneyAccount.name) == n.lower())).scalars().first()
            if not ex:
                db.add(models.MoneyAccount(type=t, name=n, currency="RUB", is_active=True))
        db.commit()

        # Categories: upsert by (type+name)
        _ensure_system_category(db, "income", "Выручка")
        _ensure_system_category(db, "expense", "Закупки")
        _ensure_system_category(db, "expense", "Аренда")
        _ensure_system_category(db, "expense", "Реклама")
        _ensure_system_category(db, "expense", "Сервис/ремонт")
        _ensure_system_category(db, "expense", "Логистика/доставка")
        _ensure_system_category(db, "expense", "Комиссии банка")
        _ensure_system_category(db, "expense", "Налоги", is_tax=True)
        _ensure_system_category(db, "expense", "Зарплата", is_payroll=True)
        _ensure_system_category(db, "income", "Выплаты маркетплейсов")
        _ensure_system_category(db, "expense", "Комиссии/услуги маркетплейсов")
        _ensure_system_category(db, "transfer", "Перевод между счетами", is_system=True)

        # Default auto-allocation rules (idempotent). User can edit/disable them.
        def _ensure_rule(
            name: str,
            match_field: str,
            pattern: str,
            direction: str,
            category_type: str,
            category_name: str,
            confidence: float,
            priority: int,
        ):
            cat = _ensure_system_category(db, category_type, category_name)
            ex = (
                db.execute(
                    select(models.MoneyRule)
                    .where(func.coalesce(models.MoneyRule.name, "") == name)
                )
                .scalars()
                .first()
            )
            if not ex:
                db.add(
                    models.MoneyRule(
                        name=name,
                        match_field=match_field,
                        pattern=pattern,
                        direction=direction,
                        category_id=cat.id,
                        confidence=confidence,
                        priority=priority,
                        is_active=True,
                    )
                )

        # Only seed if there are no rules yet
        rules_count = int(db.execute(select(func.count()).select_from(models.MoneyRule)).scalar_one())
        if rules_count == 0:
            _ensure_rule(
                name="Налоги (ключевые слова)",
                match_field="text",
                pattern="фнс|налог|усн|ндс|взнос|страх|пфр|фсс",
                direction="out",
                category_type="expense",
                category_name="Налоги",
                confidence=0.99,
                priority=1000,
            )
            _ensure_rule(
                name="Аренда (ключевые слова)",
                match_field="text",
                pattern="аренд",
                direction="out",
                category_type="expense",
                category_name="Аренда",
                confidence=0.98,
                priority=900,
            )
            _ensure_rule(
                name="Комиссии банка (ключевые слова)",
                match_field="text",
                pattern="комисс|обслуж|тариф|пакет|эквайр",
                direction="out",
                category_type="expense",
                category_name="Комиссии банка",
                confidence=0.97,
                priority=800,
            )
            _ensure_rule(
                name="Реклама (ключевые слова)",
                match_field="text",
                pattern="директ|yandex|яндекс|vk|google|meta|ads",
                direction="out",
                category_type="expense",
                category_name="Реклама",
                confidence=0.90,
                priority=700,
            )
            _ensure_rule(
                name="Логистика/доставка (ключевые слова)",
                match_field="text",
                pattern="сдэк|почта|курьер|доставка|logistic",
                direction="out",
                category_type="expense",
                category_name="Логистика/доставка",
                confidence=0.90,
                priority=650,
            )
            mp_pat = "ozon|озон|wildberries|валберис|wb|яндекс маркет|yandex market"
            _ensure_rule(
                name="Маркетплейсы: выплаты (низк. увер.)",
                match_field="text",
                pattern=mp_pat,
                direction="in",
                category_type="income",
                category_name="Выплаты маркетплейсов",
                confidence=0.70,
                priority=500,
            )
            _ensure_rule(
                name="Маркетплейсы: комиссии (низк. увер.)",
                match_field="text",
                pattern=mp_pat,
                direction="out",
                category_type="expense",
                category_name="Комиссии/услуги маркетплейсов",
                confidence=0.70,
                priority=500,
            )

        db.commit()
    except Exception:
        db.rollback()

@app.get("/materials", response_model=list[schemas.MaterialOut])
def list_materials(include_void: bool = False, db: Session = Depends(get_db)):
    q = select(models.Material).order_by(models.Material.id.asc())
    if not include_void:
        q = q.where(models.Material.is_void == False)  # noqa: E712
    return db.execute(q).scalars().all()

@app.post("/materials", response_model=schemas.MaterialOut)
def create_material(payload: schemas.MaterialCreate, db: Session = Depends(get_db)):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")

    # антидубль: если есть VOID с таким name -> реактивируем
    existing = db.execute(
        select(models.Material).where(models.Material.name == name)
    ).scalars().first()

    if existing:
        if existing.is_void:
            existing.category = payload.category
            existing.base_uom = payload.base_uom
            existing.is_lot_tracked = payload.is_lot_tracked
            existing.is_void = False
            existing.voided_at = None
            existing.void_reason = None
            db.commit()
            db.refresh(existing)
            return existing
        raise HTTPException(status_code=400, detail="Material with same name already exists")

    m = models.Material(
        name=name,
        category=payload.category,
        base_uom=payload.base_uom,
        is_lot_tracked=payload.is_lot_tracked,
        is_void=False,
        voided_at=None,
        void_reason=None,
    )
    db.add(m)
    db.flush()

    # props
    if payload.props:
        for k, v in payload.props.items():
            db.add(models.MaterialProp(material_id=m.id, key=str(k), value=str(v), value_type="str"))
    db.commit()
    db.refresh(m)
    return m

@app.put("/materials/{material_id}", response_model=schemas.MaterialOut)
def update_material(material_id: int, payload: schemas.MaterialUpdate, db: Session = Depends(get_db)):
    m = db.get(models.Material, material_id)
    if not m:
        raise HTTPException(status_code=404, detail="Material not found")
    if m.is_void:
        raise HTTPException(status_code=400, detail="Material is VOID. Reactivate by creating same name or unvoid endpoint later.")

    data = payload.model_dump(exclude_unset=True)
    if "name" in data and data["name"] is not None:
        data["name"] = data["name"].strip()

    # props отдельно
    props = data.pop("props", None)

    for k, v in data.items():
        setattr(m, k, v)

    if props is not None:
        # перезаписать props целиком
        for p in list(m.props or []):
            db.delete(p)
        if props:
            for k, v in props.items():
                db.add(models.MaterialProp(material_id=m.id, key=str(k), value=str(v), value_type="str"))

    db.commit()
    db.refresh(m)
    return m

@app.post("/materials/{material_id}/void", response_model=schemas.MaterialOut)
def void_material(material_id: int, payload: schemas.MaterialVoid, db: Session = Depends(get_db)):
    m = db.get(models.Material, material_id)
    if not m:
        raise HTTPException(status_code=404, detail="Material not found")
    if m.is_void:
        return m

    m.is_void = True
    m.voided_at = datetime.utcnow()
    m.void_reason = (payload.reason or "").strip() or None

    db.commit()
    db.refresh(m)
    return m

@app.post("/purchases", response_model=schemas.PurchaseDocOut)
def create_purchase(payload: schemas.PurchaseDocCreate, db: Session = Depends(get_db)):
    try:
        doc = crud.create_purchase(db, payload)
        db.commit()
        db.refresh(doc)
        return doc
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/purchases", response_model=list[schemas.PurchaseDocOut])
def list_purchases(db: Session = Depends(get_db)):
    return db.execute(
        select(models.PurchaseDoc).order_by(models.PurchaseDoc.id.desc())
    ).scalars().all()

@app.get("/purchases/{doc_id}", response_model=schemas.PurchaseDocOut)
def get_purchase(doc_id: int, db: Session = Depends(get_db)):
    doc = db.get(models.PurchaseDoc, doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Purchase doc not found")
    return doc

@app.post("/purchases/{doc_id}/post", response_model=schemas.PurchasePostResponse)
def post_purchase(doc_id: int, db: Session = Depends(get_db)):
    try:
        lots_created = crud.post_purchase(db, doc_id)
        db.commit()
        doc = db.get(models.PurchaseDoc, doc_id)
        return schemas.PurchasePostResponse(
            purchase_doc_id=doc_id,
            status=doc.status,
            lots_created=lots_created,
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/stock/movements", response_model=list[schemas.MovementOut])
def list_movements(material_id: int | None = None, db: Session = Depends(get_db)):
    q = select(models.LotMovement, models.Lot, models.Material).join(models.Lot, models.LotMovement.lot_id == models.Lot.id).join(models.Material, models.Lot.material_id == models.Material.id).order_by(models.LotMovement.id.desc()).limit(500)
    if material_id:
        q = q.where(models.Lot.material_id == material_id)
    rows = db.execute(q).all()
    out: list[schemas.MovementOut] = []
    for mv, lot, mat in rows:
        out.append(schemas.MovementOut(
            id=mv.id,
            lot_id=mv.lot_id,
            mv_date=mv.mv_date,
            mv_type=mv.mv_type,
            qty=float(mv.qty),
            ref_type=mv.ref_type,
            ref_id=mv.ref_id,
            material_id=mat.id,
            material_name=mat.name,
            lot_unit_cost=float(lot.unit_cost),
        ))
    return out


@app.post("/stock/writeoffs", response_model=schemas.WriteoffOut)
def create_writeoff(payload: schemas.WriteoffCreate, db: Session = Depends(get_db)):
    try:
        doc = models.WriteoffDoc(reason=payload.reason, comment=payload.comment)
        db.add(doc)
        db.flush()

        for ln in payload.lines:
            mat = db.get(models.Material, ln.material_id)
            if not mat:
                raise ValueError(f"Material not found: {ln.material_id}")
            props = crud._material_props(db, mat.id)
            qty_base = crud._convert_to_base_qty(
                base_uom=mat.base_uom,
                purchase_uom=ln.uom,
                qty=float(ln.qty),
                props=props,
                uom_factor=ln.uom_factor,
                roll_length_m=None,
            )
            if qty_base <= 0:
                raise ValueError("Количество должно быть > 0")

            db.add(models.WriteoffLine(
                writeoff_doc_id=doc.id,
                material_id=mat.id,
                qty_base=qty_base,
                base_uom=mat.base_uom,
                qty_input=ln.qty,
                uom_input=ln.uom,
            ))

            mv_type = "OUT" if payload.reason == "production" else ("SCRAP" if payload.reason == "scrap" else "ADJUST")
            allocations = fifo_allocate(db, mat.id, qty_base)
            for lot, take_qty in allocations:
                take_qty = float(take_qty)
                lot.qty_out = float(lot.qty_out) + take_qty
                db.add(models.LotMovement(
                    lot_id=lot.id,
                    mv_date=datetime.utcnow(),
                    mv_type=mv_type,
                    qty=take_qty,
                    ref_type="WRITEOFF",
                    ref_id=doc.id,
                ))

        db.commit()
        db.refresh(doc)
        # ручная сборка ответа по линиям
        return schemas.WriteoffOut(
            id=doc.id,
            doc_date=doc.doc_date,
            reason=doc.reason,
            comment=doc.comment,
            lines=[schemas.WriteoffOutLine(
                material_id=l.material_id,
                qty_input=float(l.qty_input),
                uom_input=l.uom_input,
                qty_base=float(l.qty_base),
                base_uom=l.base_uom,
            ) for l in doc.lines],
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/purchases/{doc_id}/void", response_model=schemas.PurchaseDocOut)
def void_purchase(doc_id: int, payload: schemas.PurchaseVoidRequest, db: Session = Depends(get_db)):
    try:
        crud.void_purchase(db, doc_id, payload.reason)
        db.commit()
        doc = db.get(models.PurchaseDoc, doc_id)
        return doc
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/orders")
def create_order(payload: schemas.OrderCreate, db: Session = Depends(get_db)):
    o = models.Order(order_date=payload.order_date, status="DRAFT", comment=payload.comment)
    db.add(o)
    db.flush()

    for it in payload.items:
        db.add(models.OrderItem(
            order_id=o.id,
            product_name=it.product_name,
            qty=it.qty,
            width_m=it.width_m,
            height_m=it.height_m
        ))

    db.commit()
    return {"order_id": o.id}

@app.post("/orders/{order_id}/post")
def post_order(order_id: int, payload: schemas.OrderPostRequest, db: Session = Depends(get_db)):
    order = db.get(models.Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order.status != "DRAFT":
        raise HTTPException(status_code=400, detail="Order is not in DRAFT")

    try:
        material_cost_total = 0.0
        for line in payload.consumption:
            allocations = fifo_allocate(db, line.material_id, line.qty)
            for lot, take_qty in allocations:
                take_qty = float(take_qty)
                cost = take_qty * float(lot.unit_cost)
                material_cost_total += cost

                lot.qty_out = float(lot.qty_out) + take_qty

                db.add(models.OrderConsumption(
                    order_id=order.id,
                    material_id=line.material_id,
                    lot_id=lot.id,
                    qty=take_qty,
                    uom=line.uom,
                    fifo_cost=cost
                ))

                db.add(models.LotMovement(
                    lot_id=lot.id,
                    mv_date=datetime.utcnow(),
                    mv_type="OUT",
                    qty=take_qty,
                    ref_type="ORDER",
                    ref_id=order.id
                ))

        labor_cost = 0.0
        if payload.minutes and payload.rate_rub_per_hour:
            labor_cost = (payload.minutes / 60.0) * float(payload.rate_rub_per_hour)
            if payload.employee_id:
                db.add(models.OrderLabor(
                    order_id=order.id,
                    employee_id=payload.employee_id,
                    minutes=payload.minutes,
                    rate_rub_per_hour=payload.rate_rub_per_hour,
                    labor_cost=labor_cost
                ))

        ink_ml = float(payload.c_ml + payload.m_ml + payload.y_ml + payload.k_ml)
        ink_cost = ink_ml * float(payload.ink_price_per_ml or 0)
        db.add(models.OrderInkUsage(
            order_id=order.id,
            c_ml=payload.c_ml,
            m_ml=payload.m_ml,
            y_ml=payload.y_ml,
            k_ml=payload.k_ml,
            ink_cost=ink_cost
        ))

        order.status = "OK"
        db.commit()
        return {
            "order_id": order.id,
            "status": order.status,
            "material_cost": round(material_cost_total, 2),
            "labor_cost": round(labor_cost, 2),
            "ink_cost": round(ink_cost, 2),
            "total_cost": round(material_cost_total + labor_cost + ink_cost, 2),
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/sales")
def create_sale(payload: schemas.SaleCreate, db: Session = Depends(get_db)):
    order = db.get(models.Order, payload.order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    s = models.Sale(
        sale_date=payload.sale_date,
        order_id=payload.order_id,
        marketplace=payload.marketplace,
        gross_price=payload.gross_price,
        vat_rate=payload.vat_rate
    )
    db.add(s)
    db.flush()

    for ch in payload.charges:
        db.add(models.SaleCharge(
            sale_id=s.id,
            charge_type=ch.charge_type,
            amount=ch.amount,
            comment=ch.comment
        ))

    db.commit()
    return {"sale_id": s.id}

@app.get("/orders/{order_id}/unit_economics", response_model=schemas.UnitEconomicsOut)
def unit_economics(order_id: int, db: Session = Depends(get_db)):
    order = db.get(models.Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    material_cost = db.execute(
        select(func.coalesce(func.sum(models.OrderConsumption.fifo_cost), 0)).where(models.OrderConsumption.order_id == order_id)
    ).scalar_one()

    ink_cost = db.execute(
        select(func.coalesce(models.OrderInkUsage.ink_cost, 0)).where(models.OrderInkUsage.order_id == order_id)
    ).scalar_one_or_none() or 0

    labor_cost = db.execute(
        select(func.coalesce(func.sum(models.OrderLabor.labor_cost), 0)).where(models.OrderLabor.order_id == order_id)
    ).scalar_one()

    total_cost = float(material_cost) + float(ink_cost) + float(labor_cost)

    sale = db.execute(
        select(models.Sale).where(models.Sale.order_id == order_id).order_by(models.Sale.id.desc())
    ).scalars().first()

    gross = float(sale.gross_price) if sale else 0.0
    charges_total = 0.0
    if sale:
        charges_total = float(db.execute(
            select(func.coalesce(func.sum(models.SaleCharge.amount), 0)).where(models.SaleCharge.sale_id == sale.id)
        ).scalar_one())

    net_revenue = gross - charges_total
    profit = net_revenue - total_cost

    return schemas.UnitEconomicsOut(
        order_id=order_id,
        material_cost=float(material_cost),
        ink_cost=float(ink_cost),
        labor_cost=float(labor_cost),
        total_cost=float(total_cost),
        gross_price=float(gross),
        charges_total=float(charges_total),
        net_revenue=float(net_revenue),
        profit=float(profit),
    )

@app.get("/stock/lots")
def stock_lots(db: Session = Depends(get_db)):
    """
    РЎРєР»Р°Рґ РїРѕ РїР°СЂС‚РёСЏРј: РѕСЃС‚Р°С‚РѕРє = qty_in - qty_out.
    """
    rows = db.execute(
        select(
            models.Lot.id.label("lot_id"),
            models.Material.id.label("material_id"),
            models.Material.name.label("material_name"),
            models.Material.category.label("category"),
            models.Lot.qty_in.label("qty_in"),
            models.Lot.qty_out.label("qty_out"),
            (models.Lot.qty_in - models.Lot.qty_out).label("qty_remaining"),
            models.Lot.unit_cost.label("unit_cost"),
            models.Lot.created_at.label("created_at"),
        )
        .join(models.Material, models.Material.id == models.Lot.material_id)
        .order_by(models.Lot.created_at.asc())
    ).all()

    result = []
    for r in rows:
        result.append({
            "lot_id": int(r.lot_id),
            "material_id": int(r.material_id),
            "material_name": r.material_name,
            "category": r.category,
            "qty_in": float(r.qty_in),
            "qty_out": float(r.qty_out),
            "qty_remaining": float(r.qty_remaining),
            "unit_cost": float(r.unit_cost),
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return result




# --- Biz Orders (Sales) ---
from datetime import date as _date

@app.get("/biz_orders", response_model=list[schemas.BizOrderOut])
def list_biz_orders(date_from: _date | None = None, date_to: _date | None = None, db: Session = Depends(get_db)):
    return crud.list_biz_orders(db, date_from=date_from, date_to=date_to)

@app.post("/biz_orders", response_model=schemas.BizOrderOut)
def create_biz_order(payload: schemas.BizOrderCreate, db: Session = Depends(get_db)):
    if not payload.channel:
        raise HTTPException(status_code=400, detail="channel is required")
    return crud.create_biz_order(
        db,
        order_date=payload.order_date,
        channel=payload.channel,
        subchannel=payload.subchannel,
        revenue=payload.revenue,
        comment=payload.comment,
    )

@app.patch("/biz_orders/{order_id}", response_model=schemas.BizOrderOut)
def patch_biz_order(order_id: int, payload: schemas.BizOrderUpdate, db: Session = Depends(get_db)):
    try:
        return crud.update_biz_order(db, order_id, payload.dict(exclude_unset=True))
    except ValueError:
        raise HTTPException(status_code=404, detail="order not found")

# --- Expenses (Finance) ---
@app.get("/expenses", response_model=list[schemas.ExpenseOut])
def list_expenses(date_from: _date | None = None, date_to: _date | None = None, db: Session = Depends(get_db)):
    return crud.list_expenses(db, date_from=date_from, date_to=date_to)

@app.post("/expenses", response_model=schemas.ExpenseOut)
def create_expense(payload: schemas.ExpenseCreate, db: Session = Depends(get_db)):
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="amount must be > 0")
    return crud.create_expense(
        db,
        exp_date=payload.exp_date,
        category=payload.category,
        amount=payload.amount,
        channel=payload.channel,
        comment=payload.comment,
    )

# --- Control ---
@app.get("/control", response_model=schemas.ControlOut)
def control(db: Session = Depends(get_db)):
    return crud.get_control(db)


# -----------------------------
# Money Ledger (Facts + Interpretation)
# -----------------------------

def _fingerprint(account_id: str, posted_at: datetime, amount: float, counterparty: str | None, description: str | None) -> str:
    # Stable anti-duplicate fingerprint for imports without external_id
    key = "|".join([
        str(account_id),
        posted_at.strftime("%Y-%m-%d"),
        f"{amount:.2f}",
        (counterparty or "").strip().lower(),
        (description or "").strip().lower(),
    ])
    return hashlib.sha256(key.encode("utf-8")).hexdigest()



def _decode_statement_bytes(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1251", "windows-1251"):
        try:
            return data.decode(enc)
        except Exception:
            pass
    return data.decode("latin-1", errors="replace")


def _sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[";", ",", "\t"])
        return dialect.delimiter
    except Exception:
        # fallback: prefer ';' for RU bank exports
        first = sample.splitlines()[0] if sample else ""
        return ";" if first.count(";") >= first.count(",") else ","


def _parse_decimal(val: str | None) -> float | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # remove spaces, replace comma decimal
    s = s.replace(" ", "").replace("\xa0", "").replace(",", ".")
    # keep only digits, dot, minus
    allowed = set("0123456789.-")
    s2 = "".join(ch for ch in s if ch in allowed)
    if s2 in ("", "-", ".", "-."):
        return None
    try:
        return float(s2)
    except Exception:
        return None


def _parse_dt(val: str | None) -> datetime | None:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    fmts = [
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
    ]
    for f in fmts:
        try:
            dt = datetime.strptime(s, f)
            if "%H" not in f:
                dt = dt.replace(hour=12, minute=0, second=0)
            return dt
        except Exception:
            continue
    return None


def _norm_header(h: str) -> str:
    return (
        "".join(ch for ch in h.strip().lower().replace("ё", "е") if ch.isalnum() or ch in ("_", " "))
        .replace(" ", "_")
    )


def _pick(headers: list[str], *cands: str) -> str | None:
    hs = {_norm_header(h): h for h in headers}
    for c in cands:
        key = _norm_header(c)
        if key in hs:
            return hs[key]
    for nh, orig in hs.items():
        for c in cands:
            ck = _norm_header(c)
            if ck and ck in nh:
                return orig
    return None


def _parse_bank_csv(text_csv: str) -> list[dict]:
    if not text_csv.strip():
        return []
    sample = "\n".join(text_csv.splitlines()[:20])
    delim = _sniff_delimiter(sample)
    reader = csv.DictReader(io.StringIO(text_csv), delimiter=delim)
    headers = reader.fieldnames or []

    col_date = _pick(headers, "Дата операции", "Дата", "Дата_операции", "operation_date", "date")
    col_amount = _pick(headers, "Сумма", "Amount", "сумма_операции")
    col_debit = _pick(headers, "Дебет", "Расход", "Списание", "debit")
    col_credit = _pick(headers, "Кредит", "Приход", "Зачисление", "credit")
    col_currency = _pick(headers, "Валюта", "Currency")
    col_counterparty = _pick(
        headers,
        "Контрагент",
        "Плательщик",
        "Получатель",
        "Наименование",
        "counterparty",
        "payee",
        "payer",
    )
    col_desc = _pick(headers, "Назначение платежа", "Назначение", "Описание", "Детали", "purpose", "description")
    col_ext = _pick(headers, "ID операции", "Номер операции", "Номер документа", "Уникальный идентификатор", "external_id", "id")

    out: list[dict] = []
    for row in reader:
        dt = _parse_dt(row.get(col_date) if col_date else None)
        if not dt:
            continue

        currency = (row.get(col_currency) if col_currency else None) or "RUB"
        currency = str(currency).strip().upper()[:3] if currency else "RUB"

        amount = _parse_decimal(row.get(col_amount) if col_amount else None)
        if amount is None and (col_debit or col_credit):
            debit = _parse_decimal(row.get(col_debit) if col_debit else None) or 0.0
            credit = _parse_decimal(row.get(col_credit) if col_credit else None) or 0.0
            amount = credit - debit
        if amount is None:
            continue

        counterparty = row.get(col_counterparty) if col_counterparty else None
        counterparty = str(counterparty).strip() if counterparty else None

        desc = row.get(col_desc) if col_desc else None
        desc = str(desc).strip() if desc else None

        external_id = row.get(col_ext) if col_ext else None
        external_id = str(external_id).strip() if external_id else None

        out.append(
            {
                "posted_at": dt,
                "amount": float(amount),
                "currency": currency or "RUB",
                "counterparty": counterparty,
                "description": desc,
                "external_id": external_id,
                "raw": row,
            }
        )
    return out


def _text_blob(op: "models.MoneyOperation") -> str:
    return f"{(op.counterparty or '').lower()} {(op.description or '').lower()}"


def _ensure_system_category(
    db: Session,
    typ: str,
    name: str,
    *,
    is_tax: bool = False,
    is_payroll: bool = False,
    is_system: bool = False,
) -> "models.Category":
    """Idempotent category upsert by (type + case-insensitive name)."""
    existing = db.execute(
        select(models.Category)
        .where(models.Category.type == typ)
        .where(func.lower(models.Category.name) == name.lower())
    ).scalars().first()
    if existing:
        changed = False
        if is_tax and not existing.is_tax_related:
            existing.is_tax_related = True
            changed = True
        if is_payroll and not existing.is_payroll_related:
            existing.is_payroll_related = True
            changed = True
        if is_system and not existing.is_system:
            existing.is_system = True
            changed = True
        if not existing.is_active:
            existing.is_active = True
            changed = True
        if changed:
            existing.updated_at = datetime.utcnow()
            db.commit()
        return existing

    cat = models.Category(
        name=name,
        type=typ,
        is_tax_related=is_tax,
        is_payroll_related=is_payroll,
        is_system=is_system,
        is_active=True,
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)
    db.add(
        models.AuditLog(
            entity_type="Category",
            entity_id=str(cat.id),
            action="create",
            changed_fields={"name": cat.name, "type": cat.type, "system": cat.is_system},
        )
    )
    db.commit()
    return cat


def _suggest_category_for_op(db: Session, op: "models.MoneyOperation") -> tuple[uuid.UUID, float, str] | None:
    """Rule-based suggestions.

    The main source of truth for suggestions is DB-stored MoneyRule (editable).
    Returns (category_id, confidence, note) or None.
    """
    if op.is_void:
        return None
    if op.operation_type == "transfer" or op.transfer_group_id is not None:
        return None

    amt = float(op.amount)
    direction = "in" if amt > 0 else "out" if amt < 0 else "any"

    def _pick_text(rule_field: str) -> str:
        if rule_field == "counterparty":
            return (op.counterparty or "")
        if rule_field == "description":
            return (op.description or "")
        if rule_field == "source":
            return (op.source or "")
        # default: text blob
        return _text_blob(op)

    def _match(pattern: str, txt: str) -> bool:
        t = (txt or "").lower()
        # pattern supports '|' or ',' separated keywords
        parts = [p.strip().lower() for p in re.split(r"[\|,;]+", pattern or "") if p.strip()]
        if not parts:
            return False
        return any(p in t for p in parts)

    rules = (
        db.execute(
            select(models.MoneyRule)
            .where(models.MoneyRule.is_active == True)  # noqa: E712
            .order_by(models.MoneyRule.priority.desc(), models.MoneyRule.created_at.desc())
        )
        .scalars()
        .all()
    )
    for r in rules:
        if r.account_id and r.account_id != op.account_id:
            continue
        if r.direction in ("in", "out") and direction != r.direction:
            continue
        txt = _pick_text(r.match_field)
        if _match(r.pattern, txt):
            return (r.category_id, float(r.confidence or 0.0), f"rule:{r.id}")

    return None
def _parse_sberbusiness_xlsx(data: bytes) -> list[dict]:
    """
    Parse SberBusiness XLSX statement (Выписка операций по лицевому счету).
    Returns list of dict compatible with _parse_bank_csv output:
    {posted_at, amount, currency, counterparty, description, external_id, raw}
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl is required for xlsx import: {e}")

    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row containing "Дата проводки"
    header_row = None
    max_scan = min(ws.max_row, 80)
    for r in range(1, max_scan + 1):
        for c in range(1, min(ws.max_column, 40) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and "Дата проводки" in v:
                header_row = r
                break
        if header_row:
            break

    # Fallback: known layout uses row 10
    if header_row is None:
        header_row = 10

    # Known columns for SberBusiness export
    COL_POSTED_AT = 2   # B
    COL_DEBIT_ACC = 5   # E
    COL_CREDIT_ACC = 9  # I
    COL_DEBIT_AMT = 10  # J
    COL_CREDIT_AMT = 14 # N
    COL_DOC_NO = 15     # O
    COL_PURPOSE = 21    # U

    def _cell_str(r: int, c: int) -> str | None:
        v = ws.cell(r, c).value
        if v is None:
            return None
        if isinstance(v, str):
            s = v.replace("\r", "\n").strip()
            return s if s else None
        return str(v).strip()

    def _extract_name(acc_cell: str | None) -> str | None:
        if not acc_cell:
            return None
        lines = [ln.strip() for ln in acc_cell.replace("\r", "\n").split("\n") if ln and str(ln).strip()]
        if not lines:
            return None
        # usually: [account, inn, name]
        if len(lines) >= 3:
            return " ".join(lines[2:]).strip() or lines[-1]
        return lines[-1]

    out: list[dict] = []
    r = header_row + 1
    # Some exports have a subheader row after header; skip until we see datetime/date in posted_at
    max_r = ws.max_row
    while r <= max_r:
        posted_raw = ws.cell(r, COL_POSTED_AT).value
        debit_amt_raw = ws.cell(r, COL_DEBIT_AMT).value
        credit_amt_raw = ws.cell(r, COL_CREDIT_AMT).value

        # stop if we reached totals/footer: posted_at missing and no amounts for long stretch
        if posted_raw is None and debit_amt_raw is None and credit_amt_raw is None:
            r += 1
            continue

        posted_at: datetime | None = None
        if isinstance(posted_raw, datetime):
            posted_at = posted_raw
        elif isinstance(posted_raw, date):
            posted_at = datetime.combine(posted_raw, datetime.min.time()).replace(hour=12)
        elif isinstance(posted_raw, str):
            posted_at = _parse_dt(posted_raw)

        if not posted_at:
            r += 1
            continue

        debit = float(debit_amt_raw) if isinstance(debit_amt_raw, (int, float)) else (_parse_decimal(str(debit_amt_raw)) or 0.0)
        credit = float(credit_amt_raw) if isinstance(credit_amt_raw, (int, float)) else (_parse_decimal(str(credit_amt_raw)) or 0.0)
        amount = credit - debit
        if abs(amount) < 0.0001:
            r += 1
            continue

        debit_acc = _cell_str(r, COL_DEBIT_ACC)
        credit_acc = _cell_str(r, COL_CREDIT_ACC)

        counterparty = _extract_name(credit_acc if amount < 0 else debit_acc)

        purpose = _cell_str(r, COL_PURPOSE)
        doc_no = _cell_str(r, COL_DOC_NO)

        # Build stable external id to support de-dupe even if doc numbers repeat.
        ext = None
        if doc_no:
            ext = f"sber:{doc_no}:{posted_at.strftime('%Y%m%d%H%M%S')}:{abs(amount):.2f}"
        fp_raw = {
            "posted_at": posted_at.isoformat(sep=" "),
            "debit_acc": debit_acc,
            "credit_acc": credit_acc,
            "debit": debit,
            "credit": credit,
            "doc_no": doc_no,
            "purpose": purpose,
        }

        out.append(
            {
                "posted_at": posted_at,
                "amount": float(amount),
                "currency": "RUB",
                "counterparty": counterparty,
                "description": purpose,
                "external_id": ext,
                "raw": fp_raw,
            }
        )

        r += 1

    return out

def _period_key(dt: datetime) -> str:
    return dt.strftime("%Y-%m")


def _assert_period_unlocked(db: Session, dt: datetime):
    period = _period_key(dt)
    locked = db.get(models.PeriodLock, period)
    if locked:
        raise HTTPException(status_code=409, detail=f"Период {period} закрыт. Изменения запрещены.")



@app.get("/money/accounts", response_model=list[schemas.MoneyAccountOut])
def list_money_accounts(db: Session = Depends(get_db)):
    # If user has an existing DB volume without seeded refs, ensure bootstrap.
    rows = db.execute(select(models.MoneyAccount).order_by(models.MoneyAccount.name.asc())).scalars().all()
    if not rows:
        ensure_money_bootstrap(db)
        rows = db.execute(select(models.MoneyAccount).order_by(models.MoneyAccount.name.asc())).scalars().all()
    return rows


@app.post("/money/bootstrap")
def money_bootstrap(db: Session = Depends(get_db)):
    """Idempotent bootstrap for UI convenience."""
    ensure_money_bootstrap(db)
    return {"status": "ok"}


@app.post("/money/accounts", response_model=schemas.MoneyAccountOut)
def create_money_account(payload: schemas.MoneyAccountCreate, db: Session = Depends(get_db)):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    existing = db.execute(select(models.MoneyAccount).where(models.MoneyAccount.name == name)).scalars().first()
    if existing:
        return existing
    acc = models.MoneyAccount(
        type=payload.type,
        name=name,
        currency=(payload.currency or "RUB").upper(),
        external_ref=(payload.external_ref or None),
        is_active=True,
    )
    db.add(acc)
    db.commit()
    db.refresh(acc)
    db.add(models.AuditLog(entity_type="MoneyAccount", entity_id=str(acc.id), action="create", changed_fields={"name": acc.name, "type": acc.type}))
    db.commit()
    return acc


@app.get("/money/categories", response_model=list[schemas.CategoryOut])
def list_categories(db: Session = Depends(get_db)):
    return db.execute(select(models.Category).where(models.Category.is_active == True).order_by(models.Category.type.asc(), models.Category.name.asc())).scalars().all()  # noqa: E712


@app.post("/money/categories", response_model=schemas.CategoryOut)
def create_category(payload: schemas.CategoryCreate, db: Session = Depends(get_db)):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="name is required")
    cat = models.Category(
        name=name,
        type=payload.type,
        parent_id=payload.parent_id,
        is_tax_related=payload.is_tax_related,
        is_payroll_related=payload.is_payroll_related,
        is_system=False,
        is_active=True,
    )
    db.add(cat)
    db.commit()
    db.refresh(cat)
    db.add(models.AuditLog(entity_type="Category", entity_id=str(cat.id), action="create", changed_fields={"name": cat.name, "type": cat.type}))
    db.commit()
    return cat


@app.get("/money/rules", response_model=list[schemas.MoneyRuleOut])
def list_money_rules(active_only: bool = True, db: Session = Depends(get_db)):
    q = select(models.MoneyRule).order_by(models.MoneyRule.priority.desc(), models.MoneyRule.created_at.desc())
    if active_only:
        q = q.where(models.MoneyRule.is_active == True)  # noqa: E712
    return db.execute(q).scalars().all()


@app.post("/money/rules", response_model=schemas.MoneyRuleOut)
def create_money_rule(payload: schemas.MoneyRuleCreate, db: Session = Depends(get_db)):
    pat = (payload.pattern or "").strip()
    if not pat:
        raise HTTPException(status_code=400, detail="pattern is required")
    if payload.match_field not in ("text", "counterparty", "description", "source"):
        raise HTTPException(status_code=400, detail="match_field must be text/counterparty/description/source")
    if payload.direction not in ("any", "in", "out"):
        raise HTTPException(status_code=400, detail="direction must be any/in/out")
    conf = float(payload.confidence or 0.0)
    if conf <= 0 or conf > 1:
        raise HTTPException(status_code=400, detail="confidence must be 0..1")
    cat = db.get(models.Category, payload.category_id)
    if not cat or not cat.is_active:
        raise HTTPException(status_code=400, detail="category not found")
    if payload.account_id:
        acc = db.get(models.MoneyAccount, payload.account_id)
        if not acc or not acc.is_active:
            raise HTTPException(status_code=400, detail="account not found")

    r = models.MoneyRule(
        name=(payload.name or None),
        match_field=payload.match_field,
        pattern=pat,
        direction=payload.direction,
        account_id=payload.account_id,
        category_id=payload.category_id,
        confidence=conf,
        priority=int(payload.priority or 100),
        is_active=bool(payload.is_active),
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    db.add(models.AuditLog(entity_type="MoneyRule", entity_id=str(r.id), action="create", changed_fields={"name": r.name, "match_field": r.match_field, "pattern": r.pattern, "direction": r.direction, "category_id": str(r.category_id), "confidence": float(r.confidence), "priority": r.priority, "is_active": r.is_active}))
    db.commit()
    return r


@app.patch("/money/rules/{rule_id}", response_model=schemas.MoneyRuleOut)
def patch_money_rule(rule_id: str, payload: schemas.MoneyRulePatch, db: Session = Depends(get_db)):
    try:
        rid = uuid.UUID(rule_id)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid rule_id")
    r = db.get(models.MoneyRule, rid)
    if not r:
        raise HTTPException(status_code=404, detail="rule not found")
    changed: dict = {}

    if payload.name is not None and payload.name != r.name:
        changed["name"] = {"from": r.name, "to": payload.name}
        r.name = payload.name
    if payload.match_field is not None:
        if payload.match_field not in ("text", "counterparty", "description", "source"):
            raise HTTPException(status_code=400, detail="match_field must be text/counterparty/description/source")
        if payload.match_field != r.match_field:
            changed["match_field"] = {"from": r.match_field, "to": payload.match_field}
            r.match_field = payload.match_field
    if payload.pattern is not None:
        pat = (payload.pattern or "").strip()
        if not pat:
            raise HTTPException(status_code=400, detail="pattern cannot be empty")
        if pat != r.pattern:
            changed["pattern"] = {"from": r.pattern, "to": pat}
            r.pattern = pat
    if payload.direction is not None:
        if payload.direction not in ("any", "in", "out"):
            raise HTTPException(status_code=400, detail="direction must be any/in/out")
        if payload.direction != r.direction:
            changed["direction"] = {"from": r.direction, "to": payload.direction}
            r.direction = payload.direction
    if payload.account_id is not None:
        if payload.account_id:
            acc = db.get(models.MoneyAccount, payload.account_id)
            if not acc or not acc.is_active:
                raise HTTPException(status_code=400, detail="account not found")
        if payload.account_id != r.account_id:
            changed["account_id"] = {"from": str(r.account_id) if r.account_id else None, "to": str(payload.account_id) if payload.account_id else None}
            r.account_id = payload.account_id
    if payload.category_id is not None and payload.category_id != r.category_id:
        cat = db.get(models.Category, payload.category_id)
        if not cat or not cat.is_active:
            raise HTTPException(status_code=400, detail="category not found")
        changed["category_id"] = {"from": str(r.category_id), "to": str(payload.category_id)}
        r.category_id = payload.category_id
    if payload.confidence is not None:
        conf = float(payload.confidence)
        if conf <= 0 or conf > 1:
            raise HTTPException(status_code=400, detail="confidence must be 0..1")
        if conf != float(r.confidence):
            changed["confidence"] = {"from": float(r.confidence), "to": conf}
            r.confidence = conf
    if payload.priority is not None and int(payload.priority) != int(r.priority):
        changed["priority"] = {"from": int(r.priority), "to": int(payload.priority)}
        r.priority = int(payload.priority)
    if payload.is_active is not None and bool(payload.is_active) != bool(r.is_active):
        changed["is_active"] = {"from": bool(r.is_active), "to": bool(payload.is_active)}
        r.is_active = bool(payload.is_active)

    if changed:
        r.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(r)
        db.add(models.AuditLog(entity_type="MoneyRule", entity_id=str(r.id), action="update", changed_fields=changed))
        db.commit()
    return r


@app.delete("/money/rules/{rule_id}")
def delete_money_rule(rule_id: str, db: Session = Depends(get_db)):
    try:
        rid = uuid.UUID(rule_id)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid rule_id")
    r = db.get(models.MoneyRule, rid)
    if not r:
        return {"status": "ok", "deleted": False}
    r.is_active = False
    r.updated_at = datetime.utcnow()
    db.commit()
    db.add(models.AuditLog(entity_type="MoneyRule", entity_id=str(r.id), action="deactivate", changed_fields={"is_active": False}))
    db.commit()
    return {"status": "ok", "deleted": True}

@app.get("/period-locks", response_model=list[schemas.PeriodLockOut])
def list_period_locks(db: Session = Depends(get_db)):
    return db.execute(select(models.PeriodLock).order_by(models.PeriodLock.period.asc())).scalars().all()


@app.post("/period-locks", response_model=schemas.PeriodLockOut)
def lock_period(payload: schemas.PeriodLockCreate, db: Session = Depends(get_db)):
    period = (payload.period or "").strip()
    if not period or len(period) != 7 or period[4] != "-":
        raise HTTPException(status_code=400, detail="period must be YYYY-MM")
    existing = db.get(models.PeriodLock, period)
    if existing:
        return existing
    lock = models.PeriodLock(period=period, note=payload.note, locked_by=payload.locked_by)
    db.add(lock)
    db.commit()
    db.refresh(lock)
    db.add(models.AuditLog(entity_type="PeriodLock", entity_id=period, action="lock", changed_fields={"period": period, "note": payload.note}))
    db.commit()
    return lock


@app.delete("/period-locks/{period}")
def unlock_period(period: str, db: Session = Depends(get_db)):
    lock = db.get(models.PeriodLock, period)
    if not lock:
        return {"status": "ok", "deleted": False}
    db.delete(lock)
    db.commit()
    db.add(models.AuditLog(entity_type="PeriodLock", entity_id=period, action="unlock", changed_fields={"period": period}))
    db.commit()
    return {"status": "ok", "deleted": True}



@app.get("/money/operations", response_model=list[schemas.MoneyOperationOut])
def list_money_operations(
    account_id: uuid.UUID | None = None,
    date_from: _date | None = None,
    date_to: _date | None = None,
    unallocated: bool = False,
    db: Session = Depends(get_db),
):
    q = select(models.MoneyOperation)
    if account_id:
        q = q.where(models.MoneyOperation.account_id == account_id)
    if date_from:
        q = q.where(models.MoneyOperation.posted_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.where(models.MoneyOperation.posted_at < datetime.combine(date_to, datetime.max.time()))
    q = q.where(models.MoneyOperation.is_void == False)  # noqa: E712
    q = q.order_by(models.MoneyOperation.posted_at.desc()).limit(1000)
    ops = db.execute(q).scalars().all()
    if not unallocated:
        return ops

    # filter in python: allocations sum != amount OR no allocations
    out = []
    for op in ops:
        confirmed_sum = float(
            db.execute(
                select(func.coalesce(func.sum(models.MoneyAllocation.amount_part), 0))
                .where(models.MoneyAllocation.money_operation_id == op.id)
                .where(models.MoneyAllocation.confirmed == True)  # noqa: E712
            ).scalar_one()
        )
        required = abs(float(op.amount))
        if abs(confirmed_sum - required) > 0.009:
            out.append(op)
    return out


@app.post("/imports/bank/csv", response_model=schemas.BankImportResult)
def import_bank_csv(account_id: uuid.UUID = Query(...), file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Import bank statement CSV into immutable MoneyOperation facts.

    Expected: CSV from internet banking. We try to auto-detect delimiter/encoding and common RU column names.
    Duplicates are skipped via external_id (if present) and/or hash_fingerprint.
    """
    acc = db.get(models.MoneyAccount, account_id)
    if not acc or not acc.is_active:
        raise HTTPException(status_code=404, detail="account not found")

    data = file.file.read()
    text_csv = _decode_statement_bytes(data)
    rows = _parse_bank_csv(text_csv)

    imported = 0
    skipped = 0
    errors: list[str] = []

    for r in rows:
        posted_at = r.get("posted_at")
        amount = r.get("amount")
        if not posted_at or amount is None:
            continue

        currency = (r.get("currency") or acc.currency or "RUB").upper()
        counterparty = r.get("counterparty")
        description = r.get("description")
        external_id = r.get("external_id")

        fp = _fingerprint(str(acc.id), posted_at, float(amount), counterparty, description)

        op = models.MoneyOperation(
            account_id=acc.id,
            posted_at=posted_at,
            amount=float(amount),
            currency=currency,
            counterparty=counterparty,
            description=description,
            operation_type="payment",
            external_id=external_id,
            source="bank_import",
            raw_payload={"filename": file.filename, "row": r.get("raw")},
            hash_fingerprint=fp,
            is_void=False,
        )
        db.add(op)
        try:
            db.commit()
            imported += 1
        except IntegrityError:
            db.rollback()
            skipped += 1
        except Exception as e:
            db.rollback()
            errors.append(str(e))
            if len(errors) >= 20:
                break

    if imported or skipped:
        db.add(models.AuditLog(entity_type="BankImport", entity_id=str(acc.id), action="import", changed_fields={"imported": imported, "skipped": skipped, "filename": file.filename}))
        db.commit()

    return {"imported": imported, "skipped_duplicates": skipped, "errors": errors}

@app.post("/imports/bank/xlsx", response_model=schemas.BankImportResult)
def import_bank_xlsx(account_id: uuid.UUID = Query(...), file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Import bank statement XLSX.
    Currently supports SberBusiness (Выписка операций по лицевому счету) export.
    """
    acc = db.get(models.MoneyAccount, account_id)
    if not acc or not acc.is_active:
        raise HTTPException(status_code=404, detail="account not found")

    data = file.file.read()
    rows = _parse_sberbusiness_xlsx(data)

    imported = 0
    skipped = 0
    errors: list[str] = []

    for r in rows:
        posted_at = r.get("posted_at")
        amount = r.get("amount")
        if not posted_at or amount is None:
            continue

        currency = (r.get("currency") or acc.currency or "RUB").upper()
        counterparty = r.get("counterparty")
        description = r.get("description")
        external_id = r.get("external_id")

        fp = _fingerprint(str(acc.id), posted_at, float(amount), counterparty, description)

        op = models.MoneyOperation(
            account_id=acc.id,
            posted_at=posted_at,
            amount=float(amount),
            currency=currency,
            counterparty=counterparty,
            description=description,
            operation_type="payment",
            external_id=external_id,
            source="bank_import_xlsx",
            raw_payload={"filename": file.filename, "row": r.get("raw")},
            hash_fingerprint=fp,
            is_void=False,
        )
        db.add(op)
        try:
            db.commit()
            imported += 1
        except IntegrityError:
            db.rollback()
            skipped += 1
        except Exception as e:
            db.rollback()
            errors.append(str(e))
            if len(errors) >= 20:
                break

    if imported or skipped:
        db.add(models.AuditLog(entity_type="BankImport", entity_id=str(acc.id), action="import_xlsx", changed_fields={"imported": imported, "skipped": skipped, "filename": file.filename}))
        db.commit()

    return {"imported": imported, "skipped_duplicates": skipped, "errors": errors}



@app.post("/money/operations", response_model=schemas.MoneyOperationOut)
def create_money_operation(payload: schemas.MoneyOperationCreate, db: Session = Depends(get_db)):
    # immutable fact: create only
    acc = db.get(models.MoneyAccount, payload.account_id)
    if not acc:
        raise HTTPException(status_code=404, detail="account not found")
    fp = _fingerprint(payload.account_id, payload.posted_at, payload.amount, payload.counterparty, payload.description)

    op = models.MoneyOperation(
        account_id=acc.id,
        transfer_group_id=payload.transfer_group_id,
        posted_at=payload.posted_at,
        amount=payload.amount,
        currency=(payload.currency or acc.currency or "RUB").upper(),
        counterparty=payload.counterparty,
        description=payload.description,
        operation_type=payload.operation_type or "other",
        external_id=payload.external_id,
        source=payload.source,
        raw_payload=payload.raw_payload,
        hash_fingerprint=fp,
        is_void=False,
    )
    db.add(op)
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"cannot create operation: {e}")

    db.refresh(op)
    db.add(models.AuditLog(entity_type="MoneyOperation", entity_id=str(op.id), action="create", changed_fields={"amount": float(op.amount), "posted_at": op.posted_at.isoformat(), "source": op.source}))
    db.commit()
    return op


@app.post("/money/operations/{op_id}/void")
def void_money_operation(op_id: uuid.UUID, payload: schemas.MoneyOperationVoid, db: Session = Depends(get_db)):
    op = db.get(models.MoneyOperation, op_id)
    if not op:
        raise HTTPException(status_code=404, detail="operation not found")
    _assert_period_unlocked(db, op.posted_at)
    if op.is_void:
        return {"status": "ok", "already": True}
    op.is_void = True
    op.void_reason = payload.reason or "void"
    db.commit()
    db.refresh(op)
    db.add(models.AuditLog(entity_type="MoneyOperation", entity_id=str(op.id), action="void", changed_fields={"reason": op.void_reason}))
    db.commit()
    return {"status": "ok", "already": False}


@app.post("/money/transfers")
def create_transfer(payload: schemas.MoneyTransferCreate, db: Session = Depends(get_db)):
    if payload.amount <= 0:
        raise HTTPException(status_code=400, detail="amount must be > 0")
    acc_from = db.get(models.MoneyAccount, payload.from_account_id)
    acc_to = db.get(models.MoneyAccount, payload.to_account_id)
    if not acc_from or not acc_to:
        raise HTTPException(status_code=404, detail="account not found")
    if acc_from.id == acc_to.id:
        raise HTTPException(status_code=400, detail="from/to accounts must differ")

    tg = uuid.uuid4()
    desc = payload.description or "Перевод между счетами"
    cp = payload.counterparty or "Внутренний перевод"

    op_out = models.MoneyOperation(
        account_id=acc_from.id,
        transfer_group_id=tg,
        posted_at=payload.posted_at,
        amount=-abs(payload.amount),
        currency=(payload.currency or acc_from.currency or "RUB").upper(),
        counterparty=cp,
        description=desc,
        operation_type="transfer",
        external_id=None,
        source=payload.source or "manual_other",
        raw_payload=None,
        hash_fingerprint=_fingerprint(str(acc_from.id), payload.posted_at, -abs(payload.amount), cp, desc),
        is_void=False,
    )
    op_in = models.MoneyOperation(
        account_id=acc_to.id,
        transfer_group_id=tg,
        posted_at=payload.posted_at,
        amount=abs(payload.amount),
        currency=(payload.currency or acc_to.currency or "RUB").upper(),
        counterparty=cp,
        description=desc,
        operation_type="transfer",
        external_id=None,
        source=payload.source or "manual_other",
        raw_payload=None,
        hash_fingerprint=_fingerprint(str(acc_to.id), payload.posted_at, abs(payload.amount), cp, desc),
        is_void=False,
    )
    db.add(op_out)
    db.add(op_in)
    db.flush()

    # auto allocate to system transfer category (does not affect profit)
    transfer_cat = db.execute(select(models.Category).where(models.Category.type == "transfer")).scalars().first()
    if transfer_cat:
        db.add(models.MoneyAllocation(
            money_operation_id=op_out.id,
            category_id=transfer_cat.id,
            amount_part=abs(float(op_out.amount)),
            method="system",
            confirmed=True,
            note=payload.note,
        ))
        db.add(models.MoneyAllocation(
            money_operation_id=op_in.id,
            category_id=transfer_cat.id,
            amount_part=abs(float(op_in.amount)),
            method="system",
            confirmed=True,
            note=payload.note,
        ))

    db.commit()
    db.refresh(op_out)
    db.refresh(op_in)
    db.add(models.AuditLog(entity_type="MoneyTransfer", entity_id=str(tg), action="create", changed_fields={"from": str(acc_from.id), "to": str(acc_to.id), "amount": float(payload.amount)}))
    db.commit()
    return {"transfer_group_id": str(tg), "out_operation_id": str(op_out.id), "in_operation_id": str(op_in.id)}


@app.get("/money/operations/{op_id}/allocations", response_model=list[schemas.MoneyAllocationOut])
def list_allocations(op_id: uuid.UUID, db: Session = Depends(get_db)):
    return db.execute(
        select(models.MoneyAllocation)
        .where(models.MoneyAllocation.money_operation_id == op_id)
        .order_by(models.MoneyAllocation.created_at.asc())
    ).scalars().all()



# --- Reconciliation (match facts to documents) ---
@app.get("/reconciliation/matches", response_model=list[schemas.ReconciliationMatchOut])
def list_recon_matches(status: str | None = None, db: Session = Depends(get_db)):
    q = select(models.ReconciliationMatch).order_by(models.ReconciliationMatch.created_at.desc())
    if status:
        q = q.where(models.ReconciliationMatch.status == status)
    return db.execute(q).scalars().all()


@app.post("/reconciliation/matches", response_model=schemas.ReconciliationMatchOut)
def create_recon_match(payload: schemas.ReconciliationMatchCreate, db: Session = Depends(get_db)):
    op = db.get(models.MoneyOperation, payload.money_operation_id)
    if not op:
        raise HTTPException(status_code=404, detail="operation not found")
    _assert_period_unlocked(db, op.posted_at)

    match = models.ReconciliationMatch(
        money_operation_id=op.id,
        right_type=payload.right_type,
        right_id=payload.right_id,
        method=payload.method,
        score=payload.score,
        status=payload.status,
        note=payload.note,
        confirmed_at=datetime.utcnow() if payload.status == "confirmed" else None,
    )
    db.add(match)
    db.commit()
    db.refresh(match)
    db.add(models.AuditLog(entity_type="ReconciliationMatch", entity_id=str(match.id), action="create", changed_fields={"money_operation_id": str(op.id), "right_type": payload.right_type, "right_id": payload.right_id, "status": payload.status, "method": payload.method}))
    db.commit()
    return match


@app.delete("/reconciliation/matches/{match_id}")
def delete_recon_match(match_id: uuid.UUID, db: Session = Depends(get_db)):
    m = db.get(models.ReconciliationMatch, match_id)
    if not m:
        return {"status": "ok", "deleted": False}
    op = db.get(models.MoneyOperation, m.money_operation_id)
    if op:
        _assert_period_unlocked(db, op.posted_at)
    db.add(models.AuditLog(entity_type="ReconciliationMatch", entity_id=str(m.id), action="delete", changed_fields={"money_operation_id": str(m.money_operation_id), "right_type": m.right_type, "right_id": m.right_id, "status": m.status}))
    db.delete(m)
    db.commit()
    return {"status": "ok", "deleted": True}


@app.get("/reconciliation/suggestions/{op_id}")
def recon_suggestions(op_id: uuid.UUID, db: Session = Depends(get_db)):
    op = db.get(models.MoneyOperation, op_id)
    if not op:
        raise HTTPException(status_code=404, detail="operation not found")

    # Simple rules: date window +-3 days, amount match
    target = abs(float(op.amount))
    dt = op.posted_at.date()

    def score_amt(x: float) -> float:
        d = abs(target - abs(float(x)))
        # 0 diff => 1.0, 1% diff => 0.5, >=10% => ~0
        return max(0.0, 1.0 - (d / max(1.0, target)) * 5.0)

    # Expenses (outflow)
    exp_rows = db.execute(
        select(models.Expense).where(
            models.Expense.exp_date.between(dt - timedelta(days=3), dt + timedelta(days=3))
        )
    ).scalars().all()

    # Biz orders (inflow)
    order_rows = db.execute(
        select(models.BizOrder).where(
            models.BizOrder.order_date.between(dt - timedelta(days=3), dt + timedelta(days=3))
        )
    ).scalars().all()

    # Purchases docs (gross)
    pur_docs = db.execute(
        select(models.PurchaseDoc).where(
            models.PurchaseDoc.doc_date.between(dt - timedelta(days=3), dt + timedelta(days=3))
        )
    ).scalars().all()

    suggestions = []
    if op.amount < 0:
        for e in exp_rows:
            s = score_amt(float(e.amount))
            if s > 0.35:
                suggestions.append({"type": "expense", "id": str(e.id), "date": str(e.exp_date), "amount": float(e.amount), "score": round(s, 3)})
        for p in pur_docs:
            # compute gross from lines
            lines = db.execute(select(models.PurchaseLine).where(models.PurchaseLine.doc_id == p.id)).scalars().all()
            gross = 0.0
            for l in lines:
                net = float(l.qty) * float(l.unit_price)
                vat = net * (float(l.vat_rate) / 100.0)
                gross += net + vat
            s = score_amt(gross)
            if s > 0.35:
                suggestions.append({"type": "purchase", "id": str(p.id), "date": str(p.doc_date), "amount": gross, "score": round(s, 3)})
    else:
        for o in order_rows:
            s = score_amt(float(o.revenue))
            if s > 0.35:
                suggestions.append({"type": "biz_order", "id": str(o.id), "date": str(o.order_date), "amount": float(o.revenue), "score": round(s, 3)})

    suggestions.sort(key=lambda x: x["score"], reverse=True)
    return suggestions[:20]


@app.post("/money/allocations", response_model=schemas.MoneyAllocationOut)
def create_allocation(payload: schemas.MoneyAllocationCreate, db: Session = Depends(get_db)):
    op = db.get(models.MoneyOperation, payload.money_operation_id)
    if not op or op.is_void:
        raise HTTPException(status_code=404, detail="operation not found")
    _assert_period_unlocked(db, op.posted_at)
    cat = db.get(models.Category, payload.category_id)
    if not cat or not cat.is_active:
        raise HTTPException(status_code=404, detail="category not found")

    alloc = models.MoneyAllocation(
        money_operation_id=op.id,
        category_id=cat.id,
        amount_part=abs(float(payload.amount_part)),
        linked_entity_type=payload.linked_entity_type,
        linked_entity_id=payload.linked_entity_id,
        method=payload.method,
        confidence=payload.confidence,
        confirmed=payload.confirmed,
        note=payload.note,
    )
    db.add(alloc)
    db.commit()
    db.refresh(alloc)
    db.add(models.AuditLog(entity_type="MoneyAllocation", entity_id=str(alloc.id), action="create", changed_fields={"money_operation_id": str(op.id), "category_id": str(cat.id), "amount_part": float(alloc.amount_part)}))
    db.commit()
    return alloc


@app.post("/money/allocations/{alloc_id}/confirm", response_model=schemas.MoneyAllocationOut)
def confirm_allocation(alloc_id: uuid.UUID, db: Session = Depends(get_db)):
    a = db.get(models.MoneyAllocation, alloc_id)
    if not a:
        raise HTTPException(status_code=404, detail="allocation not found")
    op = db.get(models.MoneyOperation, a.money_operation_id)
    if op:
        _assert_period_unlocked(db, op.posted_at)
    a.confirmed = True
    a.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(a)
    db.add(models.AuditLog(entity_type="MoneyAllocation", entity_id=str(a.id), action="confirm", changed_fields=None))
    db.commit()
    return a


@app.patch("/money/allocations/{alloc_id}", response_model=schemas.MoneyAllocationOut)
def patch_allocation(alloc_id: uuid.UUID, payload: schemas.MoneyAllocationPatch, db: Session = Depends(get_db)):
    a = db.get(models.MoneyAllocation, alloc_id)
    if not a:
        raise HTTPException(status_code=404, detail="allocation not found")
    op = db.get(models.MoneyOperation, a.money_operation_id)
    if op:
        _assert_period_unlocked(db, op.posted_at)

    changed = {}
    if payload.category_id is not None:
        cat = db.get(models.Category, payload.category_id)
        if not cat or not cat.is_active:
            raise HTTPException(status_code=404, detail="category not found")
        if str(a.category_id) != str(cat.id):
            changed["category_id"] = {"from": str(a.category_id), "to": str(cat.id)}
            a.category_id = cat.id

    if payload.amount_part is not None:
        new_val = abs(float(payload.amount_part))
        if new_val != float(a.amount_part):
            changed["amount_part"] = {"from": float(a.amount_part), "to": new_val}
            a.amount_part = new_val

    if payload.linked_entity_type is not None and payload.linked_entity_type != a.linked_entity_type:
        changed["linked_entity_type"] = {"from": a.linked_entity_type, "to": payload.linked_entity_type}
        a.linked_entity_type = payload.linked_entity_type

    if payload.linked_entity_id is not None and payload.linked_entity_id != a.linked_entity_id:
        changed["linked_entity_id"] = {"from": a.linked_entity_id, "to": payload.linked_entity_id}
        a.linked_entity_id = payload.linked_entity_id

    if payload.confirmed is not None and bool(payload.confirmed) != bool(a.confirmed):
        changed["confirmed"] = {"from": bool(a.confirmed), "to": bool(payload.confirmed)}
        a.confirmed = bool(payload.confirmed)

    if payload.note is not None and payload.note != a.note:
        changed["note"] = {"from": a.note, "to": payload.note}
        a.note = payload.note

    if changed:
        a.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(a)
        db.add(models.AuditLog(entity_type="MoneyAllocation", entity_id=str(a.id), action="update", changed_fields=changed))
        db.commit()
    return a


@app.delete("/money/allocations/{alloc_id}")
def delete_allocation(alloc_id: uuid.UUID, db: Session = Depends(get_db)):
    a = db.get(models.MoneyAllocation, alloc_id)
    if not a:
        return {"status": "ok", "deleted": False}
    op = db.get(models.MoneyOperation, a.money_operation_id)
    if op:
        _assert_period_unlocked(db, op.posted_at)
    payload = {"money_operation_id": str(a.money_operation_id), "category_id": str(a.category_id), "amount_part": float(a.amount_part)}
    db.add(models.AuditLog(entity_type="MoneyAllocation", entity_id=str(a.id), action="delete", changed_fields=payload))
    db.delete(a)
    db.commit()
    return {"status": "ok", "deleted": True}


@app.post("/money/auto-allocate", response_model=schemas.MoneyAutoAllocateResult)
def auto_allocate(payload: schemas.MoneyAutoAllocateParams, db: Session = Depends(get_db)):
    """Create unconfirmed allocations based on simple rules.

    Ключевое: ничего не подтверждаем автоматически. Только подсказки.
    """
    errors: list[str] = []
    scanned = suggested = updated = skipped = 0

    # Ensure categories exist (idempotent)
    ensure_money_bootstrap(db)

    q = select(models.MoneyOperation).where(models.MoneyOperation.is_void == False)
    if payload.date_from:
        q = q.where(models.MoneyOperation.posted_at >= datetime.combine(payload.date_from, datetime.min.time()))
    if payload.date_to:
        q = q.where(models.MoneyOperation.posted_at <= datetime.combine(payload.date_to, datetime.max.time()))
    q = q.order_by(models.MoneyOperation.posted_at.desc()).limit(1000)

    ops = db.execute(q).scalars().all()
    for op in ops:
        scanned += 1
        try:
            _assert_period_unlocked(db, op.posted_at)

            allocs = db.execute(select(models.MoneyAllocation).where(models.MoneyAllocation.money_operation_id == op.id)).scalars().all()
            if not payload.include_already_allocated:
                # if user already allocated/confirmed manually — don't touch
                if any(a.confirmed for a in allocs) or any(a.method == "manual" for a in allocs):
                    skipped += 1
                    continue

            # remove old rule/ai suggestions (unconfirmed), then re-suggest
            for a in allocs:
                if not a.confirmed and a.method in ("rule", "ai"):
                    db.delete(a)
            db.commit()

            sug = _suggest_category_for_op(db, op)
            if not sug:
                skipped += 1
                continue
            cat_id, conf, note = sug

            # create allocation covering full operation amount
            alloc = models.MoneyAllocation(
                money_operation_id=op.id,
                category_id=cat_id,
                amount_part=abs(float(op.amount)),
                method="rule",
                confidence=conf,
                confirmed=False,
                note=note,
            )
            db.add(alloc)
            db.commit()
            db.refresh(alloc)
            db.add(models.AuditLog(entity_type="MoneyAllocation", entity_id=str(alloc.id), action="create", changed_fields={"money_operation_id": str(op.id), "category_id": str(cat_id), "amount_part": float(alloc.amount_part), "method": "rule", "confidence": conf}))
            db.commit()

            suggested += 1
        except HTTPException:
            skipped += 1
        except Exception as e:
            db.rollback()
            errors.append(f"op {op.id}: {e}")

    return schemas.MoneyAutoAllocateResult(scanned=scanned, suggested=suggested, updated=updated, skipped=skipped, errors=errors[:50])


@app.post("/money/allocations/confirm-batch", response_model=schemas.MoneyConfirmBatchResult)
def confirm_allocations_batch(payload: schemas.MoneyConfirmBatchParams, db: Session = Depends(get_db)):
    """Confirm high-confidence allocations in bulk.

    Мы подтверждаем только те операции, где подсказки закрывают остаток полностью.
    """
    errors: list[str] = []
    confirmed = skipped = 0
    min_conf = float(payload.min_confidence or 0.95)

    # Look at recent unconfirmed allocations
    q = (
        select(models.MoneyAllocation)
        .where(models.MoneyAllocation.confirmed == False)
        .where(models.MoneyAllocation.method.in_(["rule", "ai"]))
        .order_by(models.MoneyAllocation.created_at.desc())
        .limit(2000)
    )
    allocs = db.execute(q).scalars().all()

    # group by operation
    by_op: dict[uuid.UUID, list[models.MoneyAllocation]] = {}
    for a in allocs:
        by_op.setdefault(a.money_operation_id, []).append(a)

    for op_id, alist in by_op.items():
        try:
            op = db.get(models.MoneyOperation, op_id)
            if not op or op.is_void:
                skipped += len(alist)
                continue
            _assert_period_unlocked(db, op.posted_at)

            all_allocs = db.execute(select(models.MoneyAllocation).where(models.MoneyAllocation.money_operation_id == op_id)).scalars().all()
            if any(a.method == "manual" for a in all_allocs):
                skipped += len(alist)
                continue

            required = abs(float(op.amount))
            confirmed_sum = sum(float(a.amount_part) for a in all_allocs if a.confirmed)
            remaining = required - confirmed_sum

            eligible = [a for a in all_allocs if (not a.confirmed) and a.method in ("rule", "ai") and (float(a.confidence or 0) >= min_conf)]
            elig_sum = sum(float(a.amount_part) for a in eligible)

            # confirm only if eligible allocations cover the remaining amount fully
            if abs(elig_sum - remaining) > 0.01:
                skipped += len(alist)
                continue

            for a in eligible:
                a.confirmed = True
                a.updated_at = datetime.utcnow()
                db.add(models.AuditLog(entity_type="MoneyAllocation", entity_id=str(a.id), action="confirm_batch", changed_fields={"min_confidence": min_conf}))
                confirmed += 1
            db.commit()
        except Exception as e:
            db.rollback()
            errors.append(f"op {op_id}: {e}")
            skipped += len(alist)

    return schemas.MoneyConfirmBatchResult(confirmed=confirmed, skipped=skipped, errors=errors[:50])


@app.get("/reports/cash-balance")
def report_cash_balance(db: Session = Depends(get_db)):
    # balance = sum(amount) per account (excluding void)
    q = select(models.MoneyAccount.id, models.MoneyAccount.name, func.coalesce(func.sum(models.MoneyOperation.amount), 0)).join(
        models.MoneyOperation, models.MoneyOperation.account_id == models.MoneyAccount.id, isouter=True
    ).where(
        (models.MoneyOperation.is_void == False) | (models.MoneyOperation.id == None)  # noqa: E712
    ).group_by(models.MoneyAccount.id, models.MoneyAccount.name).order_by(models.MoneyAccount.name.asc())
    rows = db.execute(q).all()
    return [{"account_id": str(i), "account_name": n, "balance": float(b)} for i, n, b in rows]


@app.get("/reports/cashflow", response_model=list[schemas.CashflowRow])
def report_cashflow(date_from: _date, date_to: _date, db: Session = Depends(get_db)):
    # daily inflow/outflow from MoneyOperation
    dt_from = datetime.combine(date_from, datetime.min.time())
    dt_to = datetime.combine(date_to, datetime.max.time())
    d = func.date(models.MoneyOperation.posted_at)
    inflow = func.coalesce(func.sum(case((models.MoneyOperation.amount > 0, models.MoneyOperation.amount), else_=0)), 0)
    outflow = func.coalesce(func.sum(case((models.MoneyOperation.amount < 0, -models.MoneyOperation.amount), else_=0)), 0)
    q = select(d.label("d"), inflow.label("inflow"), outflow.label("outflow")).where(
        models.MoneyOperation.is_void == False,
        models.MoneyOperation.posted_at >= dt_from,
        models.MoneyOperation.posted_at <= dt_to,
    ).group_by(d).order_by(d.asc())
    rows = db.execute(q).all()
    return [schemas.CashflowRow(date=r.d, inflow=float(r.inflow), outflow=float(r.outflow)) for r in rows]


@app.get("/reports/profit-cash", response_model=list[schemas.ProfitCashRow])
def report_profit_cash(date_from: _date, date_to: _date, db: Session = Depends(get_db)):
    # Profit from allocations ONLY (no approximations)
    dt_from = datetime.combine(date_from, datetime.min.time())
    dt_to = datetime.combine(date_to, datetime.max.time())
    d = func.date(models.MoneyOperation.posted_at)

    income = func.coalesce(func.sum(case((models.Category.type == "income", models.MoneyAllocation.amount_part), else_=0)), 0)
    expense = func.coalesce(func.sum(case((models.Category.type == "expense", models.MoneyAllocation.amount_part), else_=0)), 0)

    q = select(
        d.label("d"),
        income.label("income"),
        expense.label("expense"),
    ).join(models.MoneyAllocation, models.MoneyAllocation.money_operation_id == models.MoneyOperation.id).join(
        models.Category, models.Category.id == models.MoneyAllocation.category_id
    ).where(
        models.MoneyOperation.is_void == False,
        models.MoneyOperation.posted_at >= dt_from,
        models.MoneyOperation.posted_at <= dt_to,
        models.MoneyAllocation.confirmed == True,
    ).group_by(d).order_by(d.asc())
    rows = db.execute(q).all()
    out: list[schemas.ProfitCashRow] = []
    for r in rows:
        inc = float(r.income)
        exp = float(r.expense)
        out.append(schemas.ProfitCashRow(date=r.d, income=inc, expense=exp, profit=inc - exp))
    return out


@app.get("/reports/unallocated")
def report_unallocated(date_from: _date | None = None, date_to: _date | None = None, db: Session = Depends(get_db)):
    # list operations where sum(alloc) != amount
    q = select(models.MoneyOperation).where(models.MoneyOperation.is_void == False)
    if date_from:
        q = q.where(models.MoneyOperation.posted_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.where(models.MoneyOperation.posted_at <= datetime.combine(date_to, datetime.max.time()))
    q = q.order_by(models.MoneyOperation.posted_at.desc()).limit(1000)
    ops = db.execute(q).scalars().all()
    res = []
    for op in ops:
        confirmed_sum = float(
            db.execute(
                select(func.coalesce(func.sum(models.MoneyAllocation.amount_part), 0))
                .where(models.MoneyAllocation.money_operation_id == op.id)
                .where(models.MoneyAllocation.confirmed == True)  # noqa: E712
            ).scalar_one()
        )
        required = abs(float(op.amount))
        if abs(confirmed_sum - required) > 0.009:
            res.append({
                "id": str(op.id),
                "posted_at": op.posted_at.isoformat(),
                "amount": float(op.amount),
                "account_id": str(op.account_id),
                "required": required,
                "confirmed": confirmed_sum,
                "unallocated": required - confirmed_sum,
            })
    return res

# -----------------------------
# Treasury: plan items + forecast
# -----------------------------


def _last_day_of_month(year: int, month: int) -> int:
    # month: 1..12
    if month == 12:
        nxt = datetime(year + 1, 1, 1)
    else:
        nxt = datetime(year, month + 1, 1)
    return (nxt - timedelta(days=1)).day


def _plan_item_occurs_on(item: models.CashPlanItem, d: _date) -> bool:
    if not item.is_active:
        return False
    if item.start_date and d < item.start_date:
        return False
    if item.end_date and d > item.end_date:
        return False

    sch = (item.schedule or "monthly").lower()

    if sch == "once":
        return item.due_date == d

    if sch == "weekly":
        if item.weekday is None:
            return False
        return d.weekday() == int(item.weekday)

    # monthly
    dom = int(item.day_of_month or 10)
    dom = max(1, min(31, dom))
    dom_effective = min(dom, _last_day_of_month(d.year, d.month))
    return d.day == dom_effective


@app.get("/treasury/plan-items", response_model=list[schemas.CashPlanItemOut])
def list_plan_items(active_only: bool = True, db: Session = Depends(get_db)):
    q = select(models.CashPlanItem)
    if active_only:
        q = q.where(models.CashPlanItem.is_active == True)
    q = q.order_by(models.CashPlanItem.is_active.desc(), models.CashPlanItem.direction.asc(), models.CashPlanItem.name.asc())
    return list(db.scalars(q).all())


@app.post("/treasury/plan-items", response_model=schemas.CashPlanItemOut)
def create_plan_item(payload: schemas.CashPlanItemCreate, db: Session = Depends(get_db)):
    sch = (payload.schedule or "monthly").lower()
    if sch == "once" and not payload.due_date:
        raise HTTPException(400, "schedule=once требует due_date")
    if sch == "monthly" and not payload.day_of_month:
        raise HTTPException(400, "schedule=monthly требует day_of_month")
    if sch == "weekly" and payload.weekday is None:
        raise HTTPException(400, "schedule=weekly требует weekday (0=Пн..6=Вс)")

    item = models.CashPlanItem(
        name=payload.name.strip(),
        direction=payload.direction,
        amount=abs(payload.amount),
        currency=(payload.currency or "RUB").upper(),
        account_id=payload.account_id,
        category_id=payload.category_id,
        schedule=sch,
        due_date=payload.due_date,
        day_of_month=payload.day_of_month,
        weekday=payload.weekday,
        start_date=payload.start_date,
        end_date=payload.end_date,
        note=payload.note,
        is_active=payload.is_active,
        updated_at=datetime.utcnow(),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@app.patch("/treasury/plan-items/{item_id}", response_model=schemas.CashPlanItemOut)
def patch_plan_item(item_id: uuid.UUID, payload: schemas.CashPlanItemPatch, db: Session = Depends(get_db)):
    item = db.get(models.CashPlanItem, item_id)
    if not item:
        raise HTTPException(404, "plan item not found")

    # Patch only provided fields
    for field, value in payload.model_dump(exclude_unset=True).items():
        if field == "amount" and value is not None:
            value = abs(float(value))
        if field == "schedule" and value is not None:
            value = str(value).lower()
        setattr(item, field, value)

    item.updated_at = datetime.utcnow()
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@app.delete("/treasury/plan-items/{item_id}")
def delete_plan_item(item_id: uuid.UUID, db: Session = Depends(get_db)):
    item = db.get(models.CashPlanItem, item_id)
    if not item:
        raise HTTPException(404, "plan item not found")
    db.delete(item)
    db.commit()
    return {"ok": True}


@app.get("/treasury/forecast", response_model=list[schemas.CashForecastRow])
def forecast_cash(days: int = 30, date_from: _date | None = None, account_id: uuid.UUID | None = None, db: Session = Depends(get_db)):
    days = max(1, min(365, int(days or 30)))
    if date_from is None:
        date_from = datetime.utcnow().date()

    dt_from = datetime.combine(date_from, datetime.min.time())

    # balance as-of start of date_from
    op_q = select(func.coalesce(func.sum(models.MoneyOperation.amount), 0)).where(models.MoneyOperation.posted_at < dt_from)
    op_q = op_q.where(models.MoneyOperation.is_void == False)
    if account_id:
        op_q = op_q.where(models.MoneyOperation.account_id == account_id)
    start_balance = float(db.execute(op_q).scalar() or 0)

    plan_q = select(models.CashPlanItem).where(models.CashPlanItem.is_active == True)
    if account_id:
        # Если у планового платежа указан конкретный счёт — берём его. Если не указан — считаем, что влияет на общий баланс.
        plan_q = plan_q.where(or_(models.CashPlanItem.account_id == account_id, models.CashPlanItem.account_id.is_(None)))
    items = list(db.scalars(plan_q).all())

    rows: list[schemas.CashForecastRow] = []
    bal = start_balance
    for i in range(days):
        d = date_from + timedelta(days=i)
        pin = 0.0
        pout = 0.0
        for it in items:
            if _plan_item_occurs_on(it, d):
                amt = float(it.amount)
                if (it.direction or "out") == "in":
                    pin += amt
                else:
                    pout += amt
        net = pin - pout
        bal += net
        rows.append(schemas.CashForecastRow(date=d, planned_in=pin, planned_out=pout, net=net, balance=bal))

    return rows


# -----------------------------
# Marketplaces: Ozon finance (Accrual layer)
# -----------------------------

OZON_BASE_URL = "https://api-seller.ozon.ru"


def _ozon_headers(conn: models.MarketplaceConnection) -> dict:
    # Ozon Seller API auth headers
    return {
        "Client-Id": conn.client_id,
        "Api-Key": conn.api_key,
        "Content-Type": "application/json",
    }


def _ozon_date_from_to(d_from: date, d_to: date) -> tuple[str, str]:
    # Ozon often accepts ISO 8601 with Z
    f = datetime.combine(d_from, datetime.min.time()).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    t = datetime.combine(d_to, datetime.max.time()).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    return f, t


def _iter_month_chunks(d_from: date, d_to: date):
    """Split into <=31 day chunks (Ozon finance endpoints often ограничены месяцем)."""
    cur = d_from
    while cur <= d_to:
        end = min(d_to, cur + timedelta(days=30))
        yield cur, end
        cur = end + timedelta(days=1)


@app.get("/integrations/marketplaces/connections", response_model=list[schemas.MarketplaceConnectionOut])
def list_marketplace_connections(marketplace: str | None = None, db: Session = Depends(get_db)):
    q = select(models.MarketplaceConnection)
    if marketplace:
        q = q.where(models.MarketplaceConnection.marketplace == marketplace)
    q = q.order_by(models.MarketplaceConnection.created_at.desc())
    items = list(db.scalars(q).all())

    def to_out(x: models.MarketplaceConnection):
        last4 = (x.api_key or "")[-4:]
        return schemas.MarketplaceConnectionOut(
            id=x.id,
            marketplace=x.marketplace,
            name=x.name,
            client_id=x.client_id,
            api_key_last4=last4,
            note=x.note,
            is_active=bool(x.is_active),
            created_at=x.created_at,
            updated_at=x.updated_at,
        )

    return [to_out(x) for x in items]


@app.post("/integrations/marketplaces/connections", response_model=schemas.MarketplaceConnectionOut)
def create_marketplace_connection(payload: schemas.MarketplaceConnectionCreate, db: Session = Depends(get_db)):
    item = models.MarketplaceConnection(
        marketplace=payload.marketplace,
        name=payload.name,
        client_id=payload.client_id.strip(),
        api_key=payload.api_key.strip(),
        note=payload.note,
        is_active=payload.is_active,
        updated_at=datetime.utcnow(),
    )
    db.add(item)
    try:
        db.commit()
    except IntegrityError as e:
        db.rollback()
        raise HTTPException(400, f"connection already exists: {e.orig}")
    db.refresh(item)
    return schemas.MarketplaceConnectionOut(
        id=item.id,
        marketplace=item.marketplace,
        name=item.name,
        client_id=item.client_id,
        api_key_last4=(item.api_key or "")[-4:],
        note=item.note,
        is_active=bool(item.is_active),
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


@app.patch("/integrations/marketplaces/connections/{conn_id}", response_model=schemas.MarketplaceConnectionOut)
def patch_marketplace_connection(conn_id: uuid.UUID, payload: schemas.MarketplaceConnectionPatch, db: Session = Depends(get_db)):
    item = db.get(models.MarketplaceConnection, conn_id)
    if not item:
        raise HTTPException(404, "connection not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        if value is None:
            continue
        if field in ("client_id", "api_key"):
            value = str(value).strip()
        setattr(item, field, value)
    item.updated_at = datetime.utcnow()
    db.add(item)
    db.commit()
    db.refresh(item)
    return schemas.MarketplaceConnectionOut(
        id=item.id,
        marketplace=item.marketplace,
        name=item.name,
        client_id=item.client_id,
        api_key_last4=(item.api_key or "")[-4:],
        note=item.note,
        is_active=bool(item.is_active),
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


@app.post("/integrations/ozon/fetch", response_model=schemas.OzonFetchResult)
def fetch_ozon_transactions(payload: schemas.OzonFetchParams, db: Session = Depends(get_db)):
    conn = db.get(models.MarketplaceConnection, payload.connection_id)
    if not conn or conn.marketplace != "ozon":
        raise HTTPException(404, "ozon connection not found")
    if not conn.is_active:
        raise HTTPException(400, "connection is disabled")

    errors: list[str] = []
    fetched = 0
    inserted = 0
    duplicates = 0

    # chunk by month to be safe
    for d_from, d_to in _iter_month_chunks(payload.date_from, payload.date_to):
        iso_from, iso_to = _ozon_date_from_to(d_from, d_to)
        page = 1
        page_size = 1000
        while True:
            body = {
                "filter": {
                    "date": {"from": iso_from, "to": iso_to},
                    "transaction_type": "all",
                },
                "page": page,
                "page_size": page_size,
            }
            try:
                with httpx.Client(timeout=30) as client:
                    r = client.post(
                        f"{OZON_BASE_URL}/v3/finance/transaction/list",
                        headers=_ozon_headers(conn),
                        json=body,
                    )
                if r.status_code >= 400:
                    errors.append(f"ozon HTTP {r.status_code}: {r.text[:200]}")
                    break
                data = r.json() or {}
                ops = (((data.get("result") or {}).get("operations")) or [])
                if not ops:
                    break
                fetched += len(ops)

                # Build rows for batch upsert to avoid noisy duplicate-key errors in Postgres logs.
                rows = []

                def _num(v):
                    try:
                        return float(v)
                    except Exception:
                        return None

                for op in ops:
                    try:
                        op_id = str(op.get("operation_id") or op.get("id") or "").strip()
                        if not op_id:
                            continue

                        op_date_raw = op.get("operation_date") or op.get("date")
                        try:
                            op_date = datetime.fromisoformat(str(op_date_raw).replace("Z", "+00:00"))
                            # store as naive UTC for consistency
                            if op_date.tzinfo is not None:
                                op_date = op_date.astimezone(tz=None).replace(tzinfo=None)
                        except Exception:
                            op_date = datetime.utcnow()

                        posting = None
                        if isinstance(op.get("posting"), dict):
                            posting = op.get("posting", {}).get("posting_number")
                        if not posting:
                            posting = op.get("posting_number")

                        rows.append(
                            {
                                "id": uuid.uuid4(),
                                "connection_id": conn.id,
                                "operation_id": op_id,
                                "operation_date": op_date,
                                "operation_type": str(op.get("operation_type")) if op.get("operation_type") is not None else None,
                                "operation_type_name": str(op.get("operation_type_name")) if op.get("operation_type_name") is not None else None,
                                "posting_number": str(posting) if posting is not None else None,
                                "type": str(op.get("type")) if op.get("type") is not None else None,
                                "amount": _num(op.get("amount")),
                                "accruals_for_sale": _num(op.get("accruals_for_sale")),
                                "sale_commission": _num(op.get("sale_commission")),
                                "delivery_charge": _num(op.get("delivery_charge")),
                                "return_delivery_charge": _num(op.get("return_delivery_charge")),
                                "raw_payload": op,
                                "imported_at": datetime.utcnow(),
                            }
                        )
                    except Exception as e:
                        errors.append(f"parse error: {type(e).__name__}: {e}")

                if rows:
                    try:
                        if db.bind.dialect.name == "postgresql":
                            stmt = pg_insert(models.OzonTransaction.__table__).values(rows)
                            stmt = stmt.on_conflict_do_nothing(index_elements=["connection_id", "operation_id"])
                            res = db.execute(stmt)
                            db.commit()
                            ins = int(res.rowcount or 0)
                            inserted += ins
                            duplicates += max(0, len(rows) - ins)
                        else:
                            # fallback for other DBs
                            for row in rows:
                                try:
                                    tx = models.OzonTransaction(**row)
                                    db.add(tx)
                                    db.commit()
                                    inserted += 1
                                except IntegrityError:
                                    db.rollback()
                                    duplicates += 1
                    except Exception as e:
                        db.rollback()
                        errors.append(f"batch insert error: {type(e).__name__}: {e}")

                # pagination heuristic: if less than page_size, last page
                if len(ops) < page_size:
                    break
                page += 1
                if page > 200:
                    errors.append("pagination stop: too many pages")
                    break
            except Exception as e:
                errors.append(f"ozon request error: {type(e).__name__}: {e}")
                break

    return schemas.OzonFetchResult(fetched=fetched, inserted=inserted, duplicates=duplicates, errors=errors)


@app.get("/integrations/ozon/transactions", response_model=list[schemas.OzonTransactionOut])
def list_ozon_transactions(
    connection_id: uuid.UUID,
    date_from: _date | None = None,
    date_to: _date | None = None,
    limit: int = 500,
    db: Session = Depends(get_db),
):
    limit = max(1, min(5000, int(limit or 500)))
    q = select(models.OzonTransaction).where(models.OzonTransaction.connection_id == connection_id)
    if date_from:
        q = q.where(models.OzonTransaction.operation_date >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.where(models.OzonTransaction.operation_date <= datetime.combine(date_to, datetime.max.time()))
    q = q.order_by(models.OzonTransaction.operation_date.desc()).limit(limit)
    return list(db.scalars(q).all())


@app.get("/integrations/ozon/summary", response_model=schemas.OzonSummary)
def ozon_summary(
    connection_id: uuid.UUID,
    date_from: _date | None = None,
    date_to: _date | None = None,
    db: Session = Depends(get_db),
):
    q = select(
        func.count(models.OzonTransaction.id),
        func.coalesce(func.sum(models.OzonTransaction.amount), 0),
        func.coalesce(func.sum(models.OzonTransaction.accruals_for_sale), 0),
        func.coalesce(func.sum(models.OzonTransaction.sale_commission), 0),
        func.coalesce(
            func.sum(
                func.coalesce(models.OzonTransaction.delivery_charge, 0)
                + func.coalesce(models.OzonTransaction.return_delivery_charge, 0)
            ),
            0,
        ),
    ).where(models.OzonTransaction.connection_id == connection_id)
    if date_from:
        q = q.where(models.OzonTransaction.operation_date >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.where(models.OzonTransaction.operation_date <= datetime.combine(date_to, datetime.max.time()))
    row = db.execute(q).first()
    if not row:
        return schemas.OzonSummary(tx_count=0, amount_total=0, sales_total=0, commission_total=0, delivery_total=0)
    tx_count, amount_total, sales_total, commission_total, delivery_total = row
    return schemas.OzonSummary(
        tx_count=int(tx_count or 0),
        amount_total=float(amount_total or 0),
        sales_total=float(sales_total or 0),
        commission_total=float(commission_total or 0),
        delivery_total=float(delivery_total or 0),
    )


# -----------------------------
# Marketplaces: Ozon FBS/rFBS (Orders / Postings)
# -----------------------------


def _parse_iso_dt(v) -> datetime | None:
    if not v:
        return None
    s = str(v)
    try:
        # handle 'Z'
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        # store as naive UTC for simplicity
        return dt.replace(tzinfo=None)
    except Exception:
        return None


def _ozon_postings_list(
    conn: models.MarketplaceConnection,
    iso_from: str,
    iso_to: str,
    *,
    status: str | None,
    limit: int,
    offset: int,
):
    body: dict = {
        "dir": "asc",
        "filter": {"since": iso_from, "to": iso_to},
        "limit": int(limit),
        "offset": int(offset),
        "with": {
            "analytics_data": True,
            "barcodes": True,
            "financial_data": True,
            "translit": False,
        },
    }
    if status:
        body["filter"]["status"] = status

    with httpx.Client(timeout=30) as client:
        r = client.post(
            f"{OZON_BASE_URL}/v3/posting/fbs/list",
            headers=_ozon_headers(conn),
            json=body,
        )
    if r.status_code >= 400:
        raise HTTPException(r.status_code, f"ozon HTTP {r.status_code}: {r.text[:500]}")
    return r.json() or {}


def _ozon_posting_get(conn: models.MarketplaceConnection, posting_number: str):
    body = {
        "posting_number": posting_number,
        "with": {
            "analytics_data": True,
            "barcodes": True,
            "financial_data": True,
            "legal_info": False,
            "product_exemplars": False,
            "related_postings": True,
            "translit": False,
        },
    }
    with httpx.Client(timeout=30) as client:
        r = client.post(
            f"{OZON_BASE_URL}/v3/posting/fbs/get",
            headers=_ozon_headers(conn),
            json=body,
        )
    if r.status_code >= 400:
        raise HTTPException(r.status_code, f"ozon HTTP {r.status_code}: {r.text[:500]}")
    return r.json() or {}


@app.post("/integrations/ozon/fbs/postings/fetch", response_model=schemas.OzonFbsFetchResult)
def fetch_ozon_fbs_postings(payload: schemas.OzonFbsFetchParams, db: Session = Depends(get_db)):
    conn = db.get(models.MarketplaceConnection, payload.connection_id)
    if not conn or conn.marketplace != "ozon":
        raise HTTPException(404, "ozon connection not found")
    if not conn.is_active:
        raise HTTPException(400, "connection is disabled")
    if not conn.client_id or not conn.api_key:
        raise HTTPException(400, "Client-Id / Api-Key is empty")

    errors: list[str] = []
    fetched = 0
    created = 0
    updated = 0

    iso_from, iso_to = _ozon_date_from_to(payload.date_from, payload.date_to)
    limit = 100
    offset = 0
    guard = 0

    while True:
        guard += 1
        if guard > 500:
            errors.append("pagination stop: too many pages")
            break

        try:
            data = _ozon_postings_list(conn, iso_from, iso_to, status=payload.status, limit=limit, offset=offset)
        except HTTPException as e:
            errors.append(str(e.detail))
            break
        except Exception as e:
            errors.append(f"ozon request error: {type(e).__name__}: {e}")
            break

        result = data.get("result") or {}
        postings = result.get("postings") or []
        has_next = bool(result.get("has_next"))

        if not postings:
            break

        fetched += len(postings)

        for p in postings:
            try:
                posting_number = str(p.get("posting_number") or "").strip()
                if not posting_number:
                    continue

                details_payload = None
                if payload.fetch_details:
                    try:
                        details = _ozon_posting_get(conn, posting_number)
                        details_payload = (details.get("result") or None)
                    except Exception as e:
                        errors.append(f"details {posting_number}: {type(e).__name__}: {e}")

                merged_payload = {"list": p}
                if details_payload is not None:
                    merged_payload["details"] = details_payload

                status = p.get("status") or (details_payload or {}).get("status")
                substatus = p.get("substatus") or (details_payload or {}).get("substatus")
                order_id = p.get("order_id") or (details_payload or {}).get("order_id")

                created_at = _parse_iso_dt(p.get("created_at") or (details_payload or {}).get("created_at"))
                in_process_at = _parse_iso_dt(p.get("in_process_at") or (details_payload or {}).get("in_process_at"))
                shipment_date = _parse_iso_dt(p.get("shipment_date") or (details_payload or {}).get("shipment_date"))

                item = db.execute(
                    select(models.OzonPosting).where(
                        models.OzonPosting.connection_id == conn.id,
                        models.OzonPosting.posting_number == posting_number,
                    )
                ).scalar_one_or_none()

                if not item:
                    item = models.OzonPosting(
                        connection_id=conn.id,
                        posting_number=posting_number,
                        order_id=str(order_id) if order_id is not None else None,
                        status=str(status) if status is not None else None,
                        substatus=str(substatus) if substatus is not None else None,
                        created_at=created_at,
                        in_process_at=in_process_at,
                        shipment_date=shipment_date,
                        raw_payload=merged_payload,
                        imported_at=datetime.utcnow(),
                        updated_at=datetime.utcnow(),
                    )
                    db.add(item)
                    db.commit()
                    db.refresh(item)
                    created += 1
                else:
                    item.order_id = str(order_id) if order_id is not None else item.order_id
                    item.status = str(status) if status is not None else item.status
                    item.substatus = str(substatus) if substatus is not None else item.substatus
                    item.created_at = created_at or item.created_at
                    item.in_process_at = in_process_at or item.in_process_at
                    item.shipment_date = shipment_date or item.shipment_date
                    item.raw_payload = merged_payload
                    item.updated_at = datetime.utcnow()
                    db.add(item)
                    db.commit()
                    updated += 1

                # refresh items
                try:
                    db.execute(delete(models.OzonPostingItem).where(models.OzonPostingItem.posting_id == item.id))
                    db.commit()
                except Exception:
                    db.rollback()

                prod_src = (details_payload or p or {}).get("products")
                if not isinstance(prod_src, list):
                    prod_src = []
                for prod in prod_src:
                    try:
                        def _num(v):
                            try:
                                return float(v)
                            except Exception:
                                return None

                        it = models.OzonPostingItem(
                            posting_id=item.id,
                            product_id=str(prod.get("product_id")) if prod.get("product_id") is not None else None,
                            offer_id=str(prod.get("offer_id")) if prod.get("offer_id") is not None else None,
                            name=str(prod.get("name")) if prod.get("name") is not None else None,
                            sku=str(prod.get("sku")) if prod.get("sku") is not None else None,
                            quantity=int(prod.get("quantity")) if prod.get("quantity") is not None else None,
                            price=_num(prod.get("price") or prod.get("sale_price") or prod.get("payout")),
                            raw_payload=prod,
                        )
                        db.add(it)
                    except Exception as e:
                        errors.append(f"item {posting_number}: {type(e).__name__}: {e}")
                db.commit()

            except IntegrityError:
                db.rollback()
            except Exception as e:
                db.rollback()
                errors.append(f"upsert {type(e).__name__}: {e}")

        if not has_next:
            break
        offset += limit

    return schemas.OzonFbsFetchResult(fetched=fetched, created=created, updated=updated, errors=errors)


@app.get("/integrations/ozon/fbs/postings", response_model=schemas.OzonPostingsPage)
def list_ozon_fbs_postings(
    connection_id: uuid.UUID,
    date_from: date | None = None,
    date_to: date | None = None,
    status: str | None = None,
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    limit = max(1, min(200, int(limit or 50)))
    offset = max(0, int(offset or 0))

    # prefer real timeline fields, fallback to imported_at
    ts = func.coalesce(models.OzonPosting.in_process_at, models.OzonPosting.created_at, models.OzonPosting.imported_at)

    q = (
        select(models.OzonPosting)
        .options(selectinload(models.OzonPosting.items))
        .where(models.OzonPosting.connection_id == connection_id)
    )
    if status:
        q = q.where(models.OzonPosting.status == status)
    if date_from:
        q = q.where(ts >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.where(ts <= datetime.combine(date_to, datetime.max.time()))

    q = q.order_by(ts.desc()).offset(offset).limit(limit + 1)
    items = list(db.scalars(q).all())
    has_next = len(items) > limit
    items = items[:limit]

    def to_out(p: models.OzonPosting) -> schemas.OzonPostingOut:
        qty_total = 0
        items_total = 0.0
        for it in (p.items or []):
            qv = int(it.quantity or 0)
            qty_total += qv
            if it.price is not None:
                try:
                    items_total += float(it.price) * qv
                except Exception:
                    pass
        return schemas.OzonPostingOut(
            id=p.id,
            posting_number=p.posting_number,
            order_id=p.order_id,
            status=p.status,
            substatus=p.substatus,
            created_at=p.created_at,
            in_process_at=p.in_process_at,
            shipment_date=p.shipment_date,
            items=[schemas.OzonPostingItemOut.model_validate(x, from_attributes=True) for x in (p.items or [])],
            items_count=len(p.items or []),
            qty_total=qty_total,
            items_total=round(items_total, 2),
        )

    return schemas.OzonPostingsPage(
        postings=[to_out(p) for p in items],
        has_next=has_next,
        next_offset=(offset + limit) if has_next else None,
    )


@app.post("/integrations/ozon/sync_all", response_model=schemas.OzonSyncResult)
def sync_ozon_all(payload: schemas.OzonSyncParams, db: Session = Depends(get_db)):
    """One-click sync: finance transactions + FBS postings into local DB (MVP)."""

    errors: list[str] = []

    fin_res = schemas.OzonFetchResult(fetched=0, inserted=0, duplicates=0, errors=[])
    ord_res = schemas.OzonFbsFetchResult(fetched=0, created=0, updated=0, errors=[])

    try:
        fin_res = fetch_ozon_transactions(
            schemas.OzonFetchParams(
                connection_id=payload.connection_id,
                date_from=payload.date_from,
                date_to=payload.date_to,
            ),
            db,
        )
    except HTTPException as e:
        errors.append(f"finance: {e.detail}")
    except Exception as e:
        errors.append(f"finance: {type(e).__name__}: {e}")

    try:
        ord_res = fetch_ozon_fbs_postings(
            schemas.OzonFbsFetchParams(
                connection_id=payload.connection_id,
                date_from=payload.date_from,
                date_to=payload.date_to,
                status=payload.status,
                fetch_details=payload.fetch_details,
            ),
            db,
        )
    except HTTPException as e:
        errors.append(f"orders: {e.detail}")
    except Exception as e:
        errors.append(f"orders: {type(e).__name__}: {e}")

    # merge child errors too
    all_err = [*(fin_res.errors or []), *(ord_res.errors or []), *errors]
    return schemas.OzonSyncResult(finance=fin_res, orders=ord_res, errors=all_err)


@app.get("/integrations/ozon/fbs/export_ut")
def export_ozon_fbs_ut(
    connection_id: uuid.UUID,
    date_from: date,
    date_to: date,
    db: Session = Depends(get_db),
):
    # pull postings
    ts = func.coalesce(models.OzonPosting.in_process_at, models.OzonPosting.created_at, models.OzonPosting.imported_at)
    q = (
        select(models.OzonPosting)
        .options(selectinload(models.OzonPosting.items))
        .where(models.OzonPosting.connection_id == connection_id)
        .where(ts >= datetime.combine(date_from, datetime.min.time()))
        .where(ts <= datetime.combine(date_to, datetime.max.time()))
        .order_by(ts.asc())
    )
    postings = list(db.scalars(q).all())

    # finance aggregation by posting_number (optional)
    tq = select(models.OzonTransaction).where(models.OzonTransaction.connection_id == connection_id)
    tq = tq.where(models.OzonTransaction.operation_date >= datetime.combine(date_from, datetime.min.time()))
    tq = tq.where(models.OzonTransaction.operation_date <= datetime.combine(date_to, datetime.max.time()))
    txs = list(db.scalars(tq).all())

    agg: dict[str, dict] = {}
    for t in txs:
        pn = (t.posting_number or "").strip()
        if not pn:
            continue
        a = agg.setdefault(pn, {"tx": 0, "amount": 0.0, "sales": 0.0, "commission": 0.0, "delivery": 0.0, "other": 0.0})
        a["tx"] += 1
        a["amount"] += float(t.amount or 0)
        a["sales"] += float(t.accruals_for_sale or 0)
        a["commission"] += float(t.sale_commission or 0)
        a["delivery"] += float((t.delivery_charge or 0) + (t.return_delivery_charge or 0))
        # other services/adjustments not covered by split fields
        other = float(t.amount or 0) - float(t.accruals_for_sale or 0) - float(t.sale_commission or 0) - float((t.delivery_charge or 0) + (t.return_delivery_charge or 0))
        a["other"] += other

    # build CSVs
    buf_orders = io.StringIO()
    w1 = csv.writer(buf_orders, delimiter=";")
    w1.writerow(
        [
            "posting_number",
            "order_id",
            "status",
            "substatus",
            "created_at",
            "in_process_at",
            "shipment_date",
            "items_count",
            "qty_total",
            "items_total",
            "ozon_tx_count",
            "ozon_amount_total",
            "ozon_sales_total",
            "ozon_commission_total",
            "ozon_delivery_total",
            "ozon_other_total",
        ]
    )

    buf_items = io.StringIO()
    w2 = csv.writer(buf_items, delimiter=";")
    w2.writerow(["posting_number", "offer_id", "product_id", "name", "sku", "quantity", "price", "line_total"])

    for p in postings:
        qty_total = 0
        items_total = 0.0
        for it in (p.items or []):
            qv = int(it.quantity or 0)
            qty_total += qv
            price = float(it.price or 0)
            items_total += price * qv
            w2.writerow(
                [
                    p.posting_number,
                    it.offer_id or "",
                    it.product_id or "",
                    it.name or "",
                    it.sku or "",
                    qv,
                    f"{price:.2f}",
                    f"{(price * qv):.2f}",
                ]
            )

        a = agg.get(p.posting_number, {"tx": 0, "amount": 0.0, "sales": 0.0, "commission": 0.0, "delivery": 0.0, "other": 0.0})
        w1.writerow(
            [
                p.posting_number,
                p.order_id or "",
                p.status or "",
                p.substatus or "",
                (p.created_at.isoformat() if p.created_at else ""),
                (p.in_process_at.isoformat() if p.in_process_at else ""),
                (p.shipment_date.isoformat() if p.shipment_date else ""),
                len(p.items or []),
                qty_total,
                f"{items_total:.2f}",
                a["tx"],
                f"{a['amount']:.2f}",
                f"{a['sales']:.2f}",
                f"{a['commission']:.2f}",
                f"{a['delivery']:.2f}",
                f"{a['other']:.2f}",
            ]
        )

    readme = """ERP v3 • Ozon FBS export (CSV, ; separator)

Файлы:
- ozon_orders.csv: 1 строка = 1 отправление (posting_number)
- ozon_order_items.csv: товары по отправлениям

Примечание: для полной сверки с выплатами/банком и для пакета под 1С УТ используйте /integrations/ozon/ut_package.
"""

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, mode="w", compression=zipfile.ZIP_DEFLATED) as z:
        # utf-8-sig adds BOM (helps Excel/Windows; 1C can also parse it fine)
        z.writestr("ozon_orders.csv", buf_orders.getvalue().encode("utf-8-sig"))
        z.writestr("ozon_order_items.csv", buf_items.getvalue().encode("utf-8-sig"))
        z.writestr("README.txt", readme.encode("utf-8-sig"))
    zbuf.seek(0)

    filename = f"ozon_fbs_ut_{date_from.isoformat()}_{date_to.isoformat()}.zip"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(zbuf, media_type="application/zip", headers=headers)


# -----------------------------
# Ozon -> Money Ledger (marketplace balance) + payout reconciliation helpers
# -----------------------------

_OZON_PAYOUT_KEYWORDS = [
    "выплат",
    "перечисл",
    "перевод",
    "продавц",
    "ozon",
    "seller",
    "to seller",
    "payment",
    "payout",
    "sellerpayment",
    "seller_payment",
    "paymenttoseller",
    "to_seller",
    "transfer",
    "withdraw",
    "withdrawal",
]


def _ozon_guess_ledger_op_type(t: models.OzonTransaction) -> str:
    """Best-effort classification of Ozon finance operations for MoneyOperation.operation_type.

    Note: Ozon "finance operations" are NOT bank movements.
    They include accruals/withholdings inside the marketplace. Payout is the moment Ozon transfers money to the seller.
    """
    s = f"{t.operation_type or ''} {t.operation_type_name or ''}".strip().lower()
    pn = (t.posting_number or "").strip()

    # payouts usually have no posting and look like transfer/payment to seller
    if not pn:
        if any(k in s for k in _OZON_PAYOUT_KEYWORDS):
            return "payout"
        # extra heuristics for camelCase types
        if ("seller" in s and "pay" in s) or ("seller" in s and "payout" in s) or ("withdraw" in s):
            return "payout"

    amt = float(t.amount or 0)
    if pn and amt > 0:
        return "income"
    if amt < 0:
        return "fee"
    return "other"


def _get_or_create_money_account(db: Session, *, name: str, type_: str = "marketplace") -> models.MoneyAccount:
    acc = db.execute(select(models.MoneyAccount).where(models.MoneyAccount.name == name)).scalars().first()
    if acc:
        return acc
    acc = models.MoneyAccount(type=type_, name=name, currency="RUB", is_active=True)
    db.add(acc)
    db.commit()
    db.refresh(acc)
    return acc


@app.post("/integrations/ozon/to_money_ledger", response_model=schemas.OzonToLedgerResult)
def ozon_to_money_ledger(payload: schemas.OzonToLedgerParams, db: Session = Depends(get_db)):
    """Copy Ozon finance operations into Money Ledger (as marketplace balance facts).

    Это нужно для:
    - единого реестра денег (Money Ledger)
    - дальнейшего Allocation
    - сверки выплат с банком

    Важно: MoneyOperation — иммутабельные факты, поэтому используем антидубль.
    """

    conn = db.get(models.MarketplaceConnection, payload.connection_id)
    if not conn or conn.marketplace != "ozon":
        raise HTTPException(404, "ozon connection not found")

    acc = _get_or_create_money_account(db, name="Ozon баланс", type_="marketplace")

    q = select(models.OzonTransaction).where(models.OzonTransaction.connection_id == conn.id)
    q = q.where(models.OzonTransaction.operation_date >= datetime.combine(payload.date_from, datetime.min.time()))
    q = q.where(models.OzonTransaction.operation_date <= datetime.combine(payload.date_to, datetime.max.time()))
    txs = list(db.scalars(q).all())

    scanned = len(txs)
    inserted = 0
    duplicates = 0
    errors: list[str] = []

    # build batch insert
    rows = []
    for t in txs:
        try:
            ext = str(t.operation_id or "").strip()
            if not ext:
                continue
            amt = float(t.amount or 0)
            op_type = _ozon_guess_ledger_op_type(t)
            desc = (t.operation_type_name or t.operation_type or "Ozon operation")
            pn = (t.posting_number or "").strip()
            if pn:
                desc = f"{desc} • posting {pn}"

            rows.append(
                {
                    "id": uuid.uuid4(),
                    "account_id": acc.id,
                    "transfer_group_id": None,
                    "posted_at": t.operation_date,
                    "amount": amt,
                    "currency": "RUB",
                    "counterparty": "Ozon",
                    "description": desc[:240],
                    "operation_type": op_type,
                    "external_id": ext,
                    "source": "ozon_finance",
                    "raw_payload": t.raw_payload,
                    "hash_fingerprint": _fingerprint(str(acc.id), t.operation_date, amt, "Ozon", desc),
                    "is_void": False,
                    "void_reason": None,
                    "created_at": datetime.utcnow(),
                }
            )
        except Exception as e:
            errors.append(f"row build: {type(e).__name__}: {e}")

    if payload.dry_run:
        return schemas.OzonToLedgerResult(scanned=scanned, inserted=0, duplicates=0, errors=errors)

    if rows:
        try:
            if db.bind.dialect.name == "postgresql":
                stmt = pg_insert(models.MoneyOperation.__table__).values(rows)
                # unique: (source, account_id, external_id)
                stmt = stmt.on_conflict_do_nothing(index_elements=["source", "account_id", "external_id"])
                res = db.execute(stmt)
                db.commit()
                ins = int(res.rowcount or 0)
                inserted += ins
                duplicates += max(0, len(rows) - ins)
            else:
                for r in rows:
                    try:
                        db.add(models.MoneyOperation(**r))
                        db.commit()
                        inserted += 1
                    except IntegrityError:
                        db.rollback()
                        duplicates += 1
        except Exception as e:
            db.rollback()
            errors.append(f"insert: {type(e).__name__}: {e}")

    if inserted or duplicates:
        db.add(models.AuditLog(entity_type="OzonLedger", entity_id=str(conn.id), action="import", changed_fields={"scanned": scanned, "inserted": inserted, "duplicates": duplicates, "date_from": payload.date_from.isoformat(), "date_to": payload.date_to.isoformat()}))
        db.commit()

    return schemas.OzonToLedgerResult(scanned=scanned, inserted=inserted, duplicates=duplicates, errors=errors)


@app.get("/integrations/ozon/ut_package")
def export_ozon_ut_package(
    connection_id: uuid.UUID,
    date_from: date,
    date_to: date,
    db: Session = Depends(get_db),
):
    """Extended ZIP for 1C UT: orders + items + full finance ops + payout helper."""

    # reuse existing export for orders/items and posting finance agg
    # pull postings
    ts = func.coalesce(models.OzonPosting.in_process_at, models.OzonPosting.created_at, models.OzonPosting.imported_at)
    q = (
        select(models.OzonPosting)
        .options(selectinload(models.OzonPosting.items))
        .where(models.OzonPosting.connection_id == connection_id)
        .where(ts >= datetime.combine(date_from, datetime.min.time()))
        .where(ts <= datetime.combine(date_to, datetime.max.time()))
        .order_by(ts.asc())
    )
    postings = list(db.scalars(q).all())

    # finance
    tq = select(models.OzonTransaction).where(models.OzonTransaction.connection_id == connection_id)
    tq = tq.where(models.OzonTransaction.operation_date >= datetime.combine(date_from, datetime.min.time()))
    tq = tq.where(models.OzonTransaction.operation_date <= datetime.combine(date_to, datetime.max.time()))
    txs = list(db.scalars(tq).all())

    # agg by posting
    agg: dict[str, dict] = {}
    for t in txs:
        pn = (t.posting_number or "").strip()
        if not pn:
            continue
        a = agg.setdefault(pn, {"tx": 0, "amount": 0.0, "sales": 0.0, "commission": 0.0, "delivery": 0.0, "other": 0.0})
        a["tx"] += 1
        a["amount"] += float(t.amount or 0)
        a["sales"] += float(t.accruals_for_sale or 0)
        a["commission"] += float(t.sale_commission or 0)
        a["delivery"] += float((t.delivery_charge or 0) + (t.return_delivery_charge or 0))
        # other services/adjustments not covered by split fields
        other = float(t.amount or 0) - float(t.accruals_for_sale or 0) - float(t.sale_commission or 0) - float((t.delivery_charge or 0) + (t.return_delivery_charge or 0))
        a["other"] += other

    # build CSVs
    buf_orders = io.StringIO()
    w1 = csv.writer(buf_orders, delimiter=";")
    w1.writerow(
        [
            "posting_number",
            "order_id",
            "status",
            "substatus",
            "created_at",
            "in_process_at",
            "shipment_date",
            "items_count",
            "qty_total",
            "items_total",
            "ozon_tx_count",
            "ozon_amount_total",
            "ozon_sales_total",
            "ozon_commission_total",
            "ozon_delivery_total",
            "ozon_other_total",
        ]
    )

    buf_items = io.StringIO()
    w2 = csv.writer(buf_items, delimiter=";")
    w2.writerow(["posting_number", "offer_id", "product_id", "name", "sku", "quantity", "price", "line_total"])

    for p in postings:
        qty_total = 0
        items_total = 0.0
        for it in (p.items or []):
            qv = int(it.quantity or 0)
            qty_total += qv
            price = float(it.price or 0)
            items_total += price * qv
            w2.writerow(
                [
                    p.posting_number,
                    it.offer_id or "",
                    it.product_id or "",
                    it.name or "",
                    it.sku or "",
                    qv,
                    f"{price:.2f}",
                    f"{(price * qv):.2f}",
                ]
            )

        a = agg.get(p.posting_number, {"tx": 0, "amount": 0.0, "sales": 0.0, "commission": 0.0, "delivery": 0.0, "other": 0.0})
        w1.writerow(
            [
                p.posting_number,
                p.order_id or "",
                p.status or "",
                p.substatus or "",
                (p.created_at.isoformat() if p.created_at else ""),
                (p.in_process_at.isoformat() if p.in_process_at else ""),
                (p.shipment_date.isoformat() if p.shipment_date else ""),
                len(p.items or []),
                qty_total,
                f"{items_total:.2f}",
                a["tx"],
                f"{a['amount']:.2f}",
                f"{a['sales']:.2f}",
                f"{a['commission']:.2f}",
                f"{a['delivery']:.2f}",
                f"{a['other']:.2f}",
            ]
        )

    # detailed finance operations
    buf_fin = io.StringIO()
    wf = csv.writer(buf_fin, delimiter=";")
    wf.writerow(
        [
            "operation_id",
            "operation_date",
            "operation_type",
            "operation_type_name",
            "posting_number",
            "amount",
            "accruals_for_sale",
            "sale_commission",
            "delivery_charge",
            "return_delivery_charge",
        ]
    )
    for t in sorted(txs, key=lambda x: x.operation_date):
        wf.writerow(
            [
                t.operation_id,
                t.operation_date.isoformat() if t.operation_date else "",
                t.operation_type or "",
                t.operation_type_name or "",
                t.posting_number or "",
                f"{float(t.amount or 0):.2f}",
                f"{float(t.accruals_for_sale or 0):.2f}",
                f"{float(t.sale_commission or 0):.2f}",
                f"{float(t.delivery_charge or 0):.2f}",
                f"{float(t.return_delivery_charge or 0):.2f}",
            ]
        )

    # payout helper (heuristic)
    payouts = {}
    for t in txs:
        if _ozon_guess_ledger_op_type(t) != "payout":
            continue
        d = (t.operation_date.date() if t.operation_date else date_from)
        g = payouts.setdefault(d, {"amount": 0.0, "ops": []})
        g["amount"] += float(t.amount or 0)
        g["ops"].append(str(t.operation_id))

    buf_payouts = io.StringIO()
    wp = csv.writer(buf_payouts, delimiter=";")
    wp.writerow(["payout_date", "amount_marketplace", "expected_bank_in", "operation_ids"])
    for d in sorted(payouts.keys()):
        amt = float(payouts[d]["amount"])
        wp.writerow([d.isoformat(), f"{amt:.2f}", f"{abs(amt):.2f}", ",".join(payouts[d]["ops"])])



    # commission report by posting (UT helper)
    buf_comm = io.StringIO()
    wc = csv.writer(buf_comm, delimiter=";")
    wc.writerow(
        [
            "posting_number",
            "posting_date",
            "status",
            "qty_total",
            "items_total",
            "ozon_sales_total",
            "ozon_commission_total",
            "ozon_delivery_total",
            "ozon_other_total",
            "ozon_net_total",
        ]
    )
    for p in postings:
        dtp = (p.in_process_at or p.created_at or p.imported_at)
        dstr = dtp.date().isoformat() if dtp else ""
        qty_total = 0
        items_total = 0.0
        for it in (p.items or []):
            qv = int(it.quantity or 0)
            qty_total += qv
            items_total += float(it.price or 0) * qv
        a = agg.get(p.posting_number, {"tx": 0, "amount": 0.0, "sales": 0.0, "commission": 0.0, "delivery": 0.0, "other": 0.0})
        wc.writerow(
            [
                p.posting_number,
                dstr,
                (f"{p.status}/{p.substatus}" if p.substatus else (p.status or "")),
                qty_total,
                f"{items_total:.2f}",
                f"{a['sales']:.2f}",
                f"{a['commission']:.2f}",
                f"{a['delivery']:.2f}",
                f"{a['other']:.2f}",
                f"{a['amount']:.2f}",
            ]
        )

    # non-posting operations summary (payouts, acquiring, adjustments, etc.)
    nonpost: dict[str, dict] = {}
    for t in txs:
        pn = (t.posting_number or "").strip()
        if pn:
            continue
        key = (t.operation_type_name or t.operation_type or "other").strip()
        g = nonpost.setdefault(key, {"count": 0, "amount": 0.0})
        g["count"] += 1
        g["amount"] += float(t.amount or 0)

    buf_np = io.StringIO()
    wn = csv.writer(buf_np, delimiter=";")
    wn.writerow(["operation_type_name", "count", "amount_total"])
    for k, v in sorted(nonpost.items(), key=lambda kv: abs(kv[1]["amount"]), reverse=True):
        wn.writerow([k, v["count"], f"{float(v['amount']):.2f}"])
    readme = """ERP v3 • Ozon package for 1C UT (CSV, ; separator)

Файлы:
- ozon_orders.csv: 1 строка = 1 отправление (posting_number)
- ozon_order_items.csv: товары по отправлениям
- ozon_finance_operations.csv: все финансовые операции (для аудита/разбора)
- ozon_payouts.csv: эвристика выплат (для сверки с банком)
- ozon_commission_report.csv: отчет по отправлениям (свод под «Отчет комиссионера»)
- ozon_nonposting_summary.csv: свод по операциям без posting (выплаты/эквайринг/корректировки)

Важно про бухгалтерию:
1) Таблица Finance — это начисления/удержания Ozon (не банковские деньги).
2) Чтобы получить «реально пришло на р/с», делайте сверку по банку (импорт выписки) и ищите выплаты.
3) Для 1С УТ обычно формируют документы: Реализация (агент/комиссионер) + Отчет комиссионера + Услуги/комиссии.
   Этот пакет — сырьё для обработки импорта в 1С (обработку сделаем позже).
"""

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, mode="w", compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr("ozon_orders.csv", buf_orders.getvalue().encode("utf-8-sig"))
        z.writestr("ozon_order_items.csv", buf_items.getvalue().encode("utf-8-sig"))
        z.writestr("ozon_finance_operations.csv", buf_fin.getvalue().encode("utf-8-sig"))
        z.writestr("ozon_payouts.csv", buf_payouts.getvalue().encode("utf-8-sig"))
        # UT helpers
        z.writestr("ozon_commission_report.csv", buf_comm.getvalue().encode("utf-8-sig"))
        z.writestr("ozon_nonposting_summary.csv", buf_np.getvalue().encode("utf-8-sig"))
        z.writestr("README.txt", readme.encode("utf-8-sig"))
    zbuf.seek(0)

    filename = f"ozon_ut_package_{date_from.isoformat()}_{date_to.isoformat()}.zip"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(zbuf, media_type="application/zip", headers=headers)


@app.get("/integrations/ozon/payouts/reconciliation", response_model=list[schemas.OzonPayoutReconRow])
def ozon_payouts_reconciliation(
    connection_id: uuid.UUID,
    date_from: date,
    date_to: date,
    bank_account_id: uuid.UUID | None = None,
    db: Session = Depends(get_db),
):
    """Suggest matches between Ozon payouts and bank statement MoneyOperations.

    Требует импорт выписки банка в Money Ledger.
    """

    # payouts from ozon_transactions
    tq = select(models.OzonTransaction).where(models.OzonTransaction.connection_id == connection_id)
    tq = tq.where(models.OzonTransaction.operation_date >= datetime.combine(date_from, datetime.min.time()))
    tq = tq.where(models.OzonTransaction.operation_date <= datetime.combine(date_to, datetime.max.time()))
    txs = list(db.scalars(tq).all())

    groups: dict[date, dict] = {}
    for t in txs:
        if _ozon_guess_ledger_op_type(t) != "payout":
            continue
        d = (t.operation_date.date() if t.operation_date else date_from)
        g = groups.setdefault(d, {"amount": 0.0, "ops": []})
        g["amount"] += float(t.amount or 0)
        g["ops"].append(str(t.operation_id))

    # pick bank account
    bank_acc = None
    if bank_account_id:
        bank_acc = db.get(models.MoneyAccount, bank_account_id)
    if not bank_acc:
        bank_acc = db.execute(
            select(models.MoneyAccount)
            .where(models.MoneyAccount.type == "bank")
            .where(models.MoneyAccount.is_active == True)  # noqa: E712
            .order_by(models.MoneyAccount.created_at.asc())
        ).scalars().first()

    bank_ops = []
    if bank_acc:
        q = select(models.MoneyOperation).where(models.MoneyOperation.account_id == bank_acc.id)
        q = q.where(models.MoneyOperation.is_void == False)  # noqa: E712
        q = q.where(models.MoneyOperation.posted_at >= datetime.combine(date_from - timedelta(days=4), datetime.min.time()))
        q = q.where(models.MoneyOperation.posted_at <= datetime.combine(date_to + timedelta(days=4), datetime.max.time()))
        q = q.where(models.MoneyOperation.amount > 0)
        bank_ops = list(db.scalars(q).all())

    def _text_boost(op: models.MoneyOperation) -> float:
        s = f"{op.counterparty or ''} {op.description or ''}".lower()
        if "ozon" in s or "озон" in s:
            return 0.15
        return 0.0

    def _score(payout_amt: float, payout_dt: date, op: models.MoneyOperation) -> float:
        target = abs(float(payout_amt))
        amt_diff = abs(float(op.amount) - target)
        amt_score = max(0.0, 1.0 - (amt_diff / max(1.0, target)) * 6.0)
        dd = abs((op.posted_at.date() - payout_dt).days)
        date_score = max(0.0, 1.0 - (dd / 4.0))
        return round(min(1.0, 0.6 * amt_score + 0.4 * date_score + _text_boost(op)), 3)

    out: list[schemas.OzonPayoutReconRow] = []

    # existing matches
    existing = db.execute(
        select(models.ReconciliationMatch)
        .where(models.ReconciliationMatch.right_type == "ozon_payout")
        .where(models.ReconciliationMatch.status.in_(["suggested", "confirmed"]))
    ).scalars().all()
    by_right = {}
    for m in existing:
        by_right.setdefault(m.right_id, []).append(m)

    for d in sorted(groups.keys()):
        amt = float(groups[d]["amount"])
        ops = groups[d]["ops"]
        payout_key = f"ozon_payout:{connection_id}:{d.isoformat()}:{abs(amt):.2f}"

        # top suggestions
        sugg = []
        for op in bank_ops:
            sc = _score(amt, d, op)
            if sc < 0.45:
                continue
            sugg.append((sc, op))
        sugg.sort(key=lambda x: x[0], reverse=True)
        top = sugg[:3]

        # match status (if any)
        matched_bank_op_id = None
        match_status = None
        if payout_key in by_right:
            # choose confirmed first
            ms = sorted(by_right[payout_key], key=lambda x: (x.status != "confirmed", x.created_at))
            matched_bank_op_id = ms[0].money_operation_id
            match_status = ms[0].status

        out.append(
            schemas.OzonPayoutReconRow(
                payout_key=payout_key,
                payout_date=d,
                amount_marketplace=amt,
                expected_bank_in=abs(amt),
                operation_ids=ops,
                suggestions=[
                    schemas.OzonPayoutSuggestion(
                        bank_op=schemas.BankOpMini.model_validate(op, from_attributes=True),
                        score=sc,
                    )
                    for sc, op in top
                ],
                matched_bank_op_id=matched_bank_op_id,
                match_status=match_status,
            )
        )

    return out


@app.post("/integrations/ozon/payouts/reconciliation/auto_confirm", response_model=schemas.OzonPayoutAutoConfirmResult)
def ozon_payouts_auto_confirm(payload: schemas.OzonPayoutAutoConfirmParams, db: Session = Depends(get_db)):
    """Auto-confirm payout ↔ bank matches if confidence is high.

    Creates ReconciliationMatch(status=confirmed, method=auto) for rows where:
    - payout detected in Ozon finance ops
    - best bank operation suggestion score >= threshold
    - there is no existing suggested/confirmed match for this payout_key

    This is safe to run multiple times: duplicates are skipped by unique constraint.
    """

    # payouts from ozon_transactions
    tq = select(models.OzonTransaction).where(models.OzonTransaction.connection_id == payload.connection_id)
    tq = tq.where(models.OzonTransaction.operation_date >= datetime.combine(payload.date_from, datetime.min.time()))
    tq = tq.where(models.OzonTransaction.operation_date <= datetime.combine(payload.date_to, datetime.max.time()))
    txs = list(db.scalars(tq).all())

    groups: dict[date, dict] = {}
    for t in txs:
        if _ozon_guess_ledger_op_type(t) != "payout":
            continue
        d = (t.operation_date.date() if t.operation_date else payload.date_from)
        g = groups.setdefault(d, {"amount": 0.0, "ops": []})
        g["amount"] += float(t.amount or 0)
        g["ops"].append(str(t.operation_id))

    scanned = len(groups)

    # pick bank account
    bank_acc = None
    if payload.bank_account_id:
        bank_acc = db.get(models.MoneyAccount, payload.bank_account_id)
    if not bank_acc:
        bank_acc = db.execute(
            select(models.MoneyAccount)
            .where(models.MoneyAccount.type == "bank")
            .where(models.MoneyAccount.is_active == True)  # noqa: E712
            .order_by(models.MoneyAccount.created_at.asc())
        ).scalars().first()

    bank_ops = []
    if bank_acc:
        q = select(models.MoneyOperation).where(models.MoneyOperation.account_id == bank_acc.id)
        q = q.where(models.MoneyOperation.is_void == False)  # noqa: E712
        q = q.where(models.MoneyOperation.posted_at >= datetime.combine(payload.date_from - timedelta(days=4), datetime.min.time()))
        q = q.where(models.MoneyOperation.posted_at <= datetime.combine(payload.date_to + timedelta(days=4), datetime.max.time()))
        q = q.where(models.MoneyOperation.amount > 0)
        bank_ops = list(db.scalars(q).all())

    def _text_boost(op: models.MoneyOperation) -> float:
        s = f"{op.counterparty or ''} {op.description or ''}".lower()
        if "ozon" in s or "озон" in s:
            return 0.15
        return 0.0

    def _score(payout_amt: float, payout_dt: date, op: models.MoneyOperation) -> float:
        target = abs(float(payout_amt))
        amt_diff = abs(float(op.amount) - target)
        amt_score = max(0.0, 1.0 - (amt_diff / max(1.0, target)) * 6.0)
        dd = abs((op.posted_at.date() - payout_dt).days)
        date_score = max(0.0, 1.0 - (dd / 4.0))
        return round(min(1.0, 0.6 * amt_score + 0.4 * date_score + _text_boost(op)), 3)

    # existing matches
    existing = db.execute(
        select(models.ReconciliationMatch)
        .where(models.ReconciliationMatch.right_type == "ozon_payout")
        .where(models.ReconciliationMatch.status.in_(["suggested", "confirmed"]))
    ).scalars().all()

    existing_right = {m.right_id for m in existing}

    to_insert = []
    skipped_existing = 0
    skipped_locked = 0
    errors: list[str] = []

    for d in sorted(groups.keys()):
        amt = float(groups[d]["amount"])
        payout_key = f"ozon_payout:{payload.connection_id}:{d.isoformat()}:{abs(amt):.2f}"

        if payout_key in existing_right:
            skipped_existing += 1
            continue

        # pick best op by score
        best = None
        best_score = 0.0
        for op in bank_ops:
            sc = _score(amt, d, op)
            if sc > best_score:
                best_score = sc
                best = op

        if not best or best_score < float(payload.threshold or 0.85):
            continue

        try:
            _assert_period_unlocked(db, best.posted_at)
        except HTTPException:
            skipped_locked += 1
            continue
        except Exception as e:
            skipped_locked += 1
            errors.append(f"period lock check: {type(e).__name__}: {e}")
            continue

        to_insert.append(
            {
                "id": uuid.uuid4(),
                "money_operation_id": best.id,
                "right_type": "ozon_payout",
                "right_id": payout_key,
                "method": "auto",
                "score": float(best_score),
                "status": "confirmed",
                "note": f"auto_confirm score={best_score}",
                "created_at": datetime.utcnow(),
                "confirmed_at": datetime.utcnow(),
            }
        )

    confirmed = 0
    if to_insert:
        try:
            if db.bind.dialect.name == "postgresql":
                stmt = pg_insert(models.ReconciliationMatch.__table__).values(to_insert)
                stmt = stmt.on_conflict_do_nothing(index_elements=["money_operation_id", "right_type", "right_id"])
                res = db.execute(stmt)
                db.commit()
                confirmed = int(res.rowcount or 0)
            else:
                for row in to_insert:
                    try:
                        db.add(models.ReconciliationMatch(**row))
                        db.commit()
                        confirmed += 1
                    except IntegrityError:
                        db.rollback()
                    except Exception as e:
                        db.rollback()
                        errors.append(f"insert: {type(e).__name__}: {e}")
        except Exception as e:
            db.rollback()
            errors.append(f"insert batch: {type(e).__name__}: {e}")

    if confirmed:
        db.add(
            models.AuditLog(
                entity_type="OzonPayoutReconcile",
                entity_id=str(payload.connection_id),
                action="auto_confirm",
                changed_fields={
                    "date_from": payload.date_from.isoformat(),
                    "date_to": payload.date_to.isoformat(),
                    "threshold": float(payload.threshold or 0.85),
                    "confirmed": confirmed,
                },
            )
        )
        db.commit()

    return schemas.OzonPayoutAutoConfirmResult(
        scanned=scanned,
        confirmed=confirmed,
        skipped_existing=skipped_existing,
        skipped_locked=skipped_locked,
        errors=errors,
    )



@app.get("/integrations/ozon/period_status", response_model=schemas.OzonPeriodStatus)
def ozon_period_status(
    connection_id: uuid.UUID,
    date_from: date,
    date_to: date,
    bank_account_id: uuid.UUID | None = None,
    db: Session = Depends(get_db),
):
    """Period closure checklist for Ozon.

    Это управленческий статус: что загружено, что выгружено в деньги, есть ли выплаты и есть ли матчи с банком.
    """

    conn = db.get(models.MarketplaceConnection, connection_id)
    if not conn or conn.marketplace != "ozon":
        raise HTTPException(404, "ozon connection not found")

    # postings (FBS)
    ts = func.coalesce(models.OzonPosting.in_process_at, models.OzonPosting.created_at, models.OzonPosting.imported_at)
    q = (
        select(models.OzonPosting)
        .options(selectinload(models.OzonPosting.items))
        .where(models.OzonPosting.connection_id == connection_id)
        .where(ts >= datetime.combine(date_from, datetime.min.time()))
        .where(ts <= datetime.combine(date_to, datetime.max.time()))
        .order_by(ts.asc())
    )
    postings = list(db.scalars(q).all())
    postings_count = len(postings)
    items_count = 0
    items_total = 0.0
    for p0 in postings:
        for it in (p0.items or []):
            items_count += 1
            qv = int(it.quantity or 0)
            items_total += float(it.price or 0) * qv

    # finance tx
    tq = select(models.OzonTransaction).where(models.OzonTransaction.connection_id == connection_id)
    tq = tq.where(models.OzonTransaction.operation_date >= datetime.combine(date_from, datetime.min.time()))
    tq = tq.where(models.OzonTransaction.operation_date <= datetime.combine(date_to, datetime.max.time()))
    txs = list(db.scalars(tq).all())

    tx_count = len(txs)
    amount_total = sum(float(t.amount or 0) for t in txs)
    sales_total = sum(float(t.accruals_for_sale or 0) for t in txs)
    commission_total = sum(float(t.sale_commission or 0) for t in txs)
    delivery_total = sum(float((t.delivery_charge or 0) + (t.return_delivery_charge or 0)) for t in txs)

    # ledger import check (ozon_finance → MoneyOperation)
    like_pat = f"{connection_id}:%"
    ledger_ops_count = int(
        db.execute(
            select(func.count())
            .select_from(models.MoneyOperation)
            .where(models.MoneyOperation.source == "ozon_finance")
            .where(models.MoneyOperation.external_id.like(like_pat))
            .where(models.MoneyOperation.posted_at >= datetime.combine(date_from, datetime.min.time()))
            .where(models.MoneyOperation.posted_at <= datetime.combine(date_to, datetime.max.time()))
        ).scalar()
        or 0
    )

    # bank import check
    bank_acc = None
    if bank_account_id:
        bank_acc = db.get(models.MoneyAccount, bank_account_id)
    if not bank_acc:
        bank_acc = db.execute(
            select(models.MoneyAccount)
            .where(models.MoneyAccount.type == "bank")
            .where(models.MoneyAccount.is_active == True)  # noqa: E712
            .order_by(models.MoneyAccount.created_at.asc())
        ).scalars().first()

    bank_ops_count = 0
    if bank_acc:
        bank_ops_count = int(
            db.execute(
                select(func.count())
                .select_from(models.MoneyOperation)
                .where(models.MoneyOperation.account_id == bank_acc.id)
                .where(models.MoneyOperation.is_void == False)  # noqa: E712
                .where(models.MoneyOperation.posted_at >= datetime.combine(date_from, datetime.min.time()))
                .where(models.MoneyOperation.posted_at <= datetime.combine(date_to, datetime.max.time()))
            ).scalar()
            or 0
        )

    # payouts detected (heuristic on txs)
    groups: dict[date, float] = {}
    for t in txs:
        if _ozon_guess_ledger_op_type(t) != "payout":
            continue
        d = (t.operation_date.date() if t.operation_date else date_from)
        groups[d] = groups.get(d, 0.0) + float(t.amount or 0)

    payout_keys = [f"ozon_payout:{connection_id}:{d.isoformat()}:{abs(float(groups[d])):.2f}" for d in sorted(groups.keys())]
    payouts_detected = len(payout_keys)
    payout_marketplace_total = sum(abs(float(groups[d])) for d in groups.keys())

    payouts_matched = 0
    bank_matched_total = 0.0
    if payout_keys:
        matched = db.execute(
            select(models.ReconciliationMatch)
            .where(models.ReconciliationMatch.right_type == "ozon_payout")
            .where(models.ReconciliationMatch.status == "confirmed")
            .where(models.ReconciliationMatch.right_id.in_(payout_keys))
        ).scalars().all()
        matched_by_key = {}
        for m0 in matched:
            matched_by_key.setdefault(m0.right_id, m0)
        payouts_matched = len(matched_by_key)

        bank_ids = {m0.money_operation_id for m0 in matched_by_key.values()}
        if bank_ids:
            ops = db.execute(select(models.MoneyOperation).where(models.MoneyOperation.id.in_(bank_ids))).scalars().all()
            bank_matched_total = sum(float(op.amount or 0) for op in ops)

    # checks
    checks: list[schemas.OzonPeriodCheck] = []
    checks.append(
        schemas.OzonPeriodCheck(
            key="finance_loaded",
            title="Загружены начисления/удержания Ozon",
            ok=tx_count > 0,
            value=f"{tx_count} операций",
            hint="Если 0 — нажми «Синхронизировать всё» или «Загрузить из Ozon».",
        )
    )
    checks.append(
        schemas.OzonPeriodCheck(
            key="orders_loaded",
            title="Загружены заказы (FBS)",
            ok=postings_count > 0,
            value=f"{postings_count} postings",
            hint="Если 0 — открой вкладку «Заказы (FBS)» и загрузи postings.",
        )
    )
    checks.append(
        schemas.OzonPeriodCheck(
            key="ledger_import",
            title="Начисления выгружены в Деньги (Money Ledger)",
            ok=ledger_ops_count > 0,
            value=f"{ledger_ops_count} операций",
            hint="Во вкладке «Финансы» нажми «Выгрузить в Деньги».",
        )
    )
    checks.append(
        schemas.OzonPeriodCheck(
            key="bank_import",
            title="Импортирована банковская выписка",
            ok=bank_ops_count > 0,
            value=f"{bank_ops_count} операций" if bank_acc else "нет банковского счета",
            hint="Импортируй выписку в «Сверка выплат» или в разделе Деньги (CSV/XLSX).",
        )
    )

    ok_payouts = payouts_detected == 0 or payouts_matched == payouts_detected
    checks.append(
        schemas.OzonPeriodCheck(
            key="payouts_detected",
            title="Найдены выплаты Ozon в начислениях",
            ok=True,
            value=str(payouts_detected),
            hint="Если 0 — возможно в периоде не было выплат (или Ozon иначе назвал операции).",
        )
    )
    checks.append(
        schemas.OzonPeriodCheck(
            key="payouts_matched",
            title="Выплаты сверены с банком",
            ok=ok_payouts,
            value=f"{payouts_matched}/{payouts_detected}",
            hint="Открой «Сверка выплат»: импорт выписки → «Авто принять» или «Принять top1».",
        )
    )

    totals = schemas.OzonPeriodTotals(
        tx_count=tx_count,
        amount_total=float(amount_total),
        sales_total=float(sales_total),
        commission_total=float(commission_total),
        delivery_total=float(delivery_total),
        postings_count=postings_count,
        items_count=items_count,
        items_total=float(items_total),
        ledger_ops_count=int(ledger_ops_count),
        bank_ops_count=int(bank_ops_count),
        payouts_detected=int(payouts_detected),
        payouts_matched=int(payouts_matched),
        payout_marketplace_total=float(payout_marketplace_total),
        bank_matched_total=float(bank_matched_total),
    )

    return schemas.OzonPeriodStatus(
        connection_id=connection_id,
        date_from=date_from,
        date_to=date_to,
        checks=checks,
        totals=totals,
    )



# ============================================================
# Yandex Market (ymarket) integration
# ============================================================

YM_BASE = "https://api.partner.market.yandex.ru"

def _ym_headers(api_key: str) -> dict:
    return {"Api-Key": api_key, "Accept": "application/json"}

def _ym_parse_dt(val: str | None) -> datetime | None:
    if not val:
        return None
    s = str(val).strip()
    if not s:
        return None
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(s)
    except Exception:
        for fmt in ("%d-%m-%Y %H:%M:%S", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
    return None

def _ym_parse_date(val: str | None) -> date | None:
    if not val:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except Exception:
        pass
    for fmt in ("%d-%m-%Y", "%d-%m-%Y %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def _ym_fmt_ddmmyyyy(d: date) -> str:
    return d.strftime("%d-%m-%Y")

def _ym_get_connection(db: Session, connection_id: uuid.UUID) -> models.MarketplaceConnection:
    conn = db.get(models.MarketplaceConnection, connection_id)
    if not conn:
        raise HTTPException(status_code=404, detail="Marketplace connection not found")
    if (conn.marketplace or "").lower() != "ymarket":
        raise HTTPException(status_code=400, detail="Connection marketplace must be 'ymarket'")
    if not conn.api_key:
        raise HTTPException(status_code=400, detail="Api-Key is empty for this connection")
    return conn

def _ym_request_json(method: str, url: str, api_key: str, params: dict | None = None, json_body: dict | None = None) -> dict:
    with httpx.Client(timeout=60.0) as client:
        r = client.request(method, url, headers=_ym_headers(api_key), params=params, json=json_body)
        if r.status_code >= 400:
            try:
                j = r.json()
            except Exception:
                j = {"error": r.text}
            raise HTTPException(status_code=400, detail=f"YMarket API error {r.status_code}: {j}")
        try:
            return r.json()
        except Exception:
            raise HTTPException(status_code=400, detail=f"YMarket API returned non-JSON: {r.text[:200]}")

@app.get("/integrations/ymarket/campaigns", response_model=list[schemas.YMarketCampaignOut])
def ymarket_list_campaigns(connection_id: uuid.UUID, db: Session = Depends(get_db)):
    conn = _ym_get_connection(db, connection_id)
    data = _ym_request_json("GET", f"{YM_BASE}/v2/campaigns", conn.api_key)
    campaigns = []
    if isinstance(data, list):
        campaigns = data
    elif isinstance(data, dict):
        if "campaigns" in data and isinstance(data["campaigns"], list):
            campaigns = data["campaigns"]
        elif "result" in data and isinstance(data["result"], dict) and isinstance(data["result"].get("campaigns"), list):
            campaigns = data["result"]["campaigns"]
        elif "result" in data and isinstance(data["result"], list):
            campaigns = data["result"]
    out = []
    for c in campaigns or []:
        biz = c.get("business") or {}
        out.append(
            schemas.YMarketCampaignOut(
                id=int(c.get("id") or 0),
                domain=c.get("domain"),
                business_id=(biz.get("id") if isinstance(biz, dict) else None),
                business_name=(biz.get("name") if isinstance(biz, dict) else None),
                placement_type=c.get("placementType") or c.get("placement_type"),
                api_availability=c.get("apiAvailability") or c.get("api_availability"),
            )
        )
    return out

def _ym_pick_business_id(conn: models.MarketplaceConnection, campaigns: list[dict]) -> int | None:
    try:
        cid = int(str(conn.client_id or "").strip() or "0")
    except Exception:
        cid = 0
    for c in campaigns:
        if int(c.get("id") or 0) == cid:
            biz = c.get("business") or {}
            if isinstance(biz, dict) and biz.get("id"):
                return int(biz["id"])
    if campaigns:
        biz = (campaigns[0].get("business") or {})
        if isinstance(biz, dict) and biz.get("id"):
            return int(biz["id"])
    return None

@app.post("/integrations/ymarket/orders/fetch")
def ymarket_fetch_orders(params: schemas.YMarketOrdersFetchParams, db: Session = Depends(get_db)):
    conn = _ym_get_connection(db, params.connection_id)
    try:
        campaign_id = int(str(conn.client_id or "").strip())
    except Exception:
        raise HTTPException(status_code=400, detail="For YMarket connection, client_id must be Campaign ID (integer).")

    limit = max(1, min(50, int(params.limit or 50)))
    fake = bool(params.fake)
    statuses = params.statuses

    date_from = params.date_from
    date_to = params.date_to
    if date_to < date_from:
        raise HTTPException(status_code=400, detail="date_to must be >= date_from")

    imported = 0
    windows = 0

    cur = date_from
    while cur <= date_to:
        win_end = min(cur + timedelta(days=29), date_to)
        win_to_excl = win_end + timedelta(days=1)
        windows += 1

        page_token = None
        while True:
            qparams = {
                "fromDate": _ym_fmt_ddmmyyyy(cur),
                "toDate": _ym_fmt_ddmmyyyy(win_to_excl),
                "limit": limit,
                "fake": str(fake).lower(),
            }
            if statuses:
                qparams["status"] = statuses
            if page_token:
                qparams["page_token"] = page_token

            data = _ym_request_json(
                "GET",
                f"{YM_BASE}/v2/campaigns/{campaign_id}/orders",
                conn.api_key,
                params=qparams,
            )

            result = data.get("result") if isinstance(data, dict) else None
            if result is None and isinstance(data, dict) and "orders" in data:
                result = data
            orders = []
            paging = {}
            if isinstance(result, dict):
                orders = result.get("orders") or result.get("items") or []
                paging = result.get("paging") or result.get("pager") or {}
            elif isinstance(data, list):
                orders = data

            for o in orders or []:
                oid = o.get("id") or o.get("orderId") or o.get("order_id")
                if oid is None:
                    continue
                try:
                    oid_i = int(oid)
                except Exception:
                    continue

                status = o.get("status")
                substatus = o.get("substatus") or o.get("subStatus")
                created_at = _ym_parse_dt(o.get("creationDate") or o.get("createdAt") or o.get("creation_date"))
                updated_at = _ym_parse_dt(o.get("updateDate") or o.get("updatedAt") or o.get("updated_at"))
                shipment_date = _ym_parse_date(o.get("shipmentDate") or o.get("supplierShipmentDate") or ((o.get("delivery") or {}).get("shipmentDate") if isinstance(o.get("delivery"), dict) else None))

                buyer_total = o.get("buyerTotal") or o.get("buyerTotalBeforeDiscount") or o.get("total") or (((o.get("delivery") or {}).get("buyerTotal")) if isinstance(o.get("delivery"), dict) else None)
                items_total = o.get("buyerItemsTotal") or o.get("buyerItemsTotalBeforeDiscount") or o.get("itemsTotal") or None
                currency = o.get("currency") if isinstance(o.get("currency"), str) else None

                stmt = (
                    pg_insert(models.YMarketOrder)
                    .values(
                        id=uuid.uuid4(),
                        connection_id=params.connection_id,
                        order_id=oid_i,
                        status=status,
                        substatus=substatus,
                        created_at=created_at,
                        updated_at=updated_at,
                        shipment_date=shipment_date,
                        buyer_total=buyer_total,
                        items_total=items_total,
                        currency=currency,
                        raw_payload=o,
                        imported_at=datetime.utcnow(),
                    )
                    .on_conflict_do_update(
                        index_elements=[models.YMarketOrder.connection_id, models.YMarketOrder.order_id],
                        set_={
                            "status": status,
                            "substatus": substatus,
                            "created_at": created_at,
                            "updated_at": updated_at,
                            "shipment_date": shipment_date,
                            "buyer_total": buyer_total,
                            "items_total": items_total,
                            "currency": currency,
                            "raw_payload": o,
                            "imported_at": datetime.utcnow(),
                        },
                    )
                    .returning(models.YMarketOrder.id)
                )
                order_row_id = db.execute(stmt).scalar_one()
                db.execute(delete(models.YMarketOrderItem).where(models.YMarketOrderItem.ymarket_order_id == order_row_id))

                items = o.get("items") or []
                for it in items or []:
                    qv = it.get("count") or it.get("quantity") or it.get("qty") or 0
                    try:
                        qv_i = int(qv)
                    except Exception:
                        qv_i = 0
                    price = it.get("buyerPrice") or it.get("price") or it.get("priceBeforeDiscount") or it.get("unitPrice") or 0
                    try:
                        price_f = float(price)
                    except Exception:
                        price_f = 0.0
                    line_total = it.get("buyerPriceTotal") or it.get("total") or (price_f * qv_i)
                    try:
                        line_total_f = float(line_total)
                    except Exception:
                        line_total_f = price_f * qv_i

                    db.add(
                        models.YMarketOrderItem(
                            ymarket_order_id=order_row_id,
                            offer_id=str(it.get("offerId") or it.get("offer_id") or "") or None,
                            shop_sku=str(it.get("shopSku") or it.get("shop_sku") or "") or None,
                            market_sku=str(it.get("marketSku") or it.get("market_sku") or "") or None,
                            name=it.get("offerName") or it.get("name"),
                            quantity=qv_i,
                            price=price_f,
                            line_total=line_total_f,
                            raw_payload=it,
                        )
                    )

                imported += 1

            db.commit()

            next_token = None
            if isinstance(paging, dict):
                next_token = paging.get("nextPageToken") or paging.get("nextPageToken".lower())
            if not next_token and isinstance(result, dict):
                next_token = result.get("nextPageToken")
            if not next_token:
                break
            page_token = next_token

        cur = win_end + timedelta(days=1)

    return {"imported": imported, "windows": windows}

@app.get("/integrations/ymarket/orders", response_model=list[schemas.YMarketOrderOut])
def ymarket_list_orders(
    connection_id: uuid.UUID,
    date_from: date | None = None,
    date_to: date | None = None,
    limit: int = 200,
    db: Session = Depends(get_db),
):
    limit = max(1, min(2000, int(limit or 200)))
    q = (
        select(models.YMarketOrder)
        .options(selectinload(models.YMarketOrder.items))
        .where(models.YMarketOrder.connection_id == connection_id)
        .order_by(models.YMarketOrder.created_at.desc().nullslast(), models.YMarketOrder.imported_at.desc())
        .limit(limit)
    )
    if date_from:
        q = q.where(models.YMarketOrder.created_at >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.where(models.YMarketOrder.created_at <= datetime.combine(date_to, datetime.max.time()))
    rows = list(db.scalars(q).all())
    out: list[schemas.YMarketOrderOut] = []
    for r in rows:
        out.append(
            schemas.YMarketOrderOut(
                id=r.id,
                connection_id=r.connection_id,
                order_id=r.order_id,
                status=r.status,
                substatus=r.substatus,
                created_at=r.created_at,
                updated_at=r.updated_at,
                shipment_date=r.shipment_date,
                buyer_total=(float(r.buyer_total) if r.buyer_total is not None else None),
                items_total=(float(r.items_total) if r.items_total is not None else None),
                currency=r.currency,
                imported_at=r.imported_at,
                items=[
                    schemas.YMarketOrderItemOut(
                        offer_id=it.offer_id,
                        shop_sku=it.shop_sku,
                        market_sku=it.market_sku,
                        name=it.name,
                        quantity=it.quantity,
                        price=(float(it.price) if it.price is not None else None),
                        line_total=(float(it.line_total) if it.line_total is not None else None),
                    )
                    for it in (r.items or [])
                ],
            )
        )
    return out

@app.get("/integrations/ymarket/reports", response_model=list[schemas.YMarketReportOut])
def ymarket_list_reports(connection_id: uuid.UUID, db: Session = Depends(get_db)):
    q = select(models.YMarketReport).where(models.YMarketReport.connection_id == connection_id).order_by(models.YMarketReport.created_at.desc())
    return list(db.scalars(q).all())

@app.post("/integrations/ymarket/reports/united-netting/generate", response_model=schemas.YMarketReportOut)
def ymarket_generate_united_netting(params: schemas.YMarketReportGenerateParams, db: Session = Depends(get_db)):
    conn = _ym_get_connection(db, params.connection_id)
    try:
        campaign_id = int(str(conn.client_id or "").strip())
    except Exception:
        raise HTTPException(status_code=400, detail="For YMarket connection, client_id must be Campaign ID (integer).")

    data = _ym_request_json("GET", f"{YM_BASE}/v2/campaigns", conn.api_key)
    campaigns = []
    if isinstance(data, list):
        campaigns = data
    elif isinstance(data, dict):
        if isinstance(data.get("result"), dict) and isinstance(data["result"].get("campaigns"), list):
            campaigns = data["result"]["campaigns"]
        elif isinstance(data.get("campaigns"), list):
            campaigns = data["campaigns"]
        elif isinstance(data.get("result"), list):
            campaigns = data["result"]
    business_id = _ym_pick_business_id(conn, campaigns)
    if not business_id:
        raise HTTPException(status_code=400, detail="Cannot determine businessId from GET /v2/campaigns. Check Api-Key.")

    if params.date_to < params.date_from:
        raise HTTPException(status_code=400, detail="date_to must be >= date_from")

    # Partner API expects `format` in (FILE|CSV|JSON). FILE is an XLSX spreadsheet.
    # UI may send legacy "XLSX" -> map to FILE.
    fmt = (params.format or "FILE").upper()
    if fmt in ("XLSX",):
        fmt = "FILE"
    if fmt not in ("FILE", "CSV", "JSON"):
        fmt = "FILE"
    lang = (params.language or "RU").upper()

    body = {
        "businessId": business_id,
        "campaignIds": [campaign_id],
        "placementPrograms": params.placement_programs or ["FBS"],
        "dateFrom": params.date_from.isoformat(),
        "dateTo": params.date_to.isoformat(),
    }

    data = _ym_request_json(
        "POST",
        f"{YM_BASE}/v2/reports/united-netting/generate",
        conn.api_key,
        params={"format": fmt, "language": lang},
        json_body=body,
    )

    report_id = None
    status = None
    if isinstance(data, dict):
        res = data.get("result") if isinstance(data.get("result"), dict) else data
        report_id = res.get("reportId") or res.get("report_id")
        status = res.get("status")
    if not report_id:
        raise HTTPException(status_code=400, detail=f"Unexpected generate response: {data}")

    stmt = (
        pg_insert(models.YMarketReport)
        .values(
            id=uuid.uuid4(),
            connection_id=params.connection_id,
            report_id=str(report_id),
            report_type="united_netting",
            status=status or "NEW",
            file_url=None,
            date_from=params.date_from,
            date_to=params.date_to,
            raw_payload=data,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        .on_conflict_do_update(
            index_elements=[models.YMarketReport.connection_id, models.YMarketReport.report_id],
            set_={
                "status": status or "NEW",
                "raw_payload": data,
                "date_from": params.date_from,
                "date_to": params.date_to,
                "updated_at": datetime.utcnow(),
            },
        )
        .returning(models.YMarketReport.id)
    )
    db.execute(stmt).scalar_one()
    db.commit()

    rep = db.execute(
        select(models.YMarketReport)
        .where(models.YMarketReport.connection_id == params.connection_id)
        .where(models.YMarketReport.report_id == str(report_id))
    ).scalar_one()
    return rep

@app.get("/integrations/ymarket/reports/info", response_model=schemas.YMarketReportInfoOut)
def ymarket_report_info(connection_id: uuid.UUID, report_id: str, db: Session = Depends(get_db)):
    conn = _ym_get_connection(db, connection_id)
    data = _ym_request_json("GET", f"{YM_BASE}/v2/reports/info/{report_id}", conn.api_key)

    res = data.get("result") if isinstance(data, dict) and isinstance(data.get("result"), dict) else (data if isinstance(data, dict) else {})
    status = res.get("status")
    file_url = res.get("file")

    rep = db.execute(
        select(models.YMarketReport)
        .where(models.YMarketReport.connection_id == connection_id)
        .where(models.YMarketReport.report_id == str(report_id))
    ).scalar_one_or_none()
    if rep:
        rep.status = status
        rep.file_url = file_url
        rep.raw_payload = data
        rep.updated_at = datetime.utcnow()
        db.commit()

    return schemas.YMarketReportInfoOut(report_id=str(report_id), status=status, file_url=file_url, raw=data if isinstance(data, dict) else None)

@app.get("/integrations/ymarket/reports/download")
def ymarket_report_download(connection_id: uuid.UUID, report_id: str, db: Session = Depends(get_db)):
    info = ymarket_report_info(connection_id=connection_id, report_id=report_id, db=db)
    if not info.file_url:
        raise HTTPException(status_code=400, detail=f"Report is not ready yet. Status: {info.status}")

    conn = _ym_get_connection(db, connection_id)

    def gen():
        with httpx.Client(timeout=120.0) as client:
            r = client.get(info.file_url, headers=_ym_headers(conn.api_key))
            if r.status_code >= 400:
                raise HTTPException(status_code=400, detail=f"Download error {r.status_code}: {r.text[:200]}")
            for chunk in r.iter_bytes(chunk_size=1024 * 512):
                if chunk:
                    yield chunk

    ext = "dat"
    lower = str(info.file_url).lower()
    if ".csv" in lower:
        ext = "csv"
    elif ".xlsx" in lower:
        ext = "xlsx"
    elif ".zip" in lower:
        ext = "zip"
    filename = f"ymarket_{report_id}.{ext}"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(gen(), media_type="application/octet-stream", headers=headers)

@app.get("/integrations/ymarket/ut_package")
def ymarket_export_ut_package(
    connection_id: uuid.UUID,
    date_from: date,
    date_to: date,
    report_id: str | None = None,
    db: Session = Depends(get_db),
):
    q = (
        select(models.YMarketOrder)
        .options(selectinload(models.YMarketOrder.items))
        .where(models.YMarketOrder.connection_id == connection_id)
        .where(models.YMarketOrder.created_at >= datetime.combine(date_from, datetime.min.time()))
        .where(models.YMarketOrder.created_at <= datetime.combine(date_to, datetime.max.time()))
        .order_by(models.YMarketOrder.created_at.asc().nullslast(), models.YMarketOrder.order_id.asc())
    )
    orders = list(db.scalars(q).all())

    rq = (
        select(models.YMarketReport)
        .where(models.YMarketReport.connection_id == connection_id)
        .where(models.YMarketReport.created_at >= datetime.combine(date_from, datetime.min.time()))
        .where(models.YMarketReport.created_at <= datetime.combine(date_to, datetime.max.time()))
        .order_by(models.YMarketReport.created_at.asc())
    )
    reports = list(db.scalars(rq).all())

    buf_orders = io.StringIO()
    w1 = csv.writer(buf_orders, delimiter=";")
    w1.writerow(
        ["order_id", "status", "substatus", "created_at", "updated_at", "shipment_date", "buyer_total", "items_total", "currency", "items_count", "qty_total"]
    )

    buf_items = io.StringIO()
    w2 = csv.writer(buf_items, delimiter=";")
    w2.writerow(["order_id", "offer_id", "shop_sku", "market_sku", "name", "quantity", "price", "line_total"])

    for o in orders:
        items_count = len(o.items or [])
        qty_total = 0
        for it in (o.items or []):
            qv = int(it.quantity or 0)
            qty_total += qv
            w2.writerow(
                [
                    o.order_id,
                    it.offer_id or "",
                    it.shop_sku or "",
                    it.market_sku or "",
                    it.name or "",
                    qv,
                    (f"{float(it.price):.2f}" if it.price is not None else ""),
                    (f"{float(it.line_total):.2f}" if it.line_total is not None else ""),
                ]
            )

        w1.writerow(
            [
                o.order_id,
                o.status or "",
                o.substatus or "",
                (o.created_at.isoformat() if o.created_at else ""),
                (o.updated_at.isoformat() if o.updated_at else ""),
                (o.shipment_date.isoformat() if o.shipment_date else ""),
                (f"{float(o.buyer_total):.2f}" if o.buyer_total is not None else ""),
                (f"{float(o.items_total):.2f}" if o.items_total is not None else ""),
                o.currency or "",
                items_count,
                qty_total,
            ]
        )

    buf_reports = io.StringIO()
    w3 = csv.writer(buf_reports, delimiter=";")
    w3.writerow(["report_id", "report_type", "status", "date_from", "date_to", "created_at", "file_url"])
    for r in reports:
        w3.writerow(
            [
                r.report_id,
                r.report_type,
                r.status or "",
                (r.date_from.isoformat() if r.date_from else ""),
                (r.date_to.isoformat() if r.date_to else ""),
                (r.created_at.isoformat() if r.created_at else ""),
                (r.file_url or ""),
            ]
        )

    readme = """ERP v3 • Yandex Market export (CSV, ; separator)

Файлы:
- ymarket_orders.csv: 1 строка = 1 заказ
- ymarket_order_items.csv: строки товаров по заказам
- ymarket_reports.csv: метаданные сформированных отчётов (например United Netting)

Подсказки:
- Для загрузки отчёта United Netting сформируйте его во вкладке «Отчёты» (генерация может занять время).
- Исторические заказы (>30 дней после доставки/отмены) API v2 может не вернуть — позже добавим получение через business orders (v1).
"""

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, mode="w", compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr("ymarket_orders.csv", buf_orders.getvalue().encode("utf-8-sig"))
        z.writestr("ymarket_order_items.csv", buf_items.getvalue().encode("utf-8-sig"))
        z.writestr("ymarket_reports.csv", buf_reports.getvalue().encode("utf-8-sig"))
        z.writestr("README.txt", readme.encode("utf-8-sig"))

        if report_id:
            try:
                info = ymarket_report_info(connection_id=connection_id, report_id=report_id, db=db)
                if info.file_url:
                    conn = _ym_get_connection(db, connection_id)
                    with httpx.Client(timeout=120.0) as client:
                        r = client.get(info.file_url, headers=_ym_headers(conn.api_key))
                        if r.status_code < 400:
                            ext = "dat"
                            lower = str(info.file_url).lower()
                            if ".csv" in lower:
                                ext = "csv"
                            elif ".xlsx" in lower:
                                ext = "xlsx"
                            elif ".zip" in lower:
                                ext = "zip"
                            z.writestr(f"united_netting_{report_id}.{ext}", r.content)
            except Exception:
                pass

    zbuf.seek(0)
    filename = f"ymarket_ut_{date_from.isoformat()}_{date_to.isoformat()}.zip"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(zbuf, media_type="application/zip", headers=headers)


# ============================================================
# Wildberries (wb) integration (Statistics API)
# ============================================================

WB_STATS_BASE = "https://statistics-api.wildberries.ru"


def _wb_headers(token: str) -> dict:
    # WB expects the token value in Authorization header as-is
    return {"Authorization": token, "Accept": "application/json"}


def _wb_parse_dt(val: str | None) -> datetime | None:
    if not val:
        return None
    s = str(val).strip()
    if not s:
        return None
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(s)
    except Exception:
        # common WB formats
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(s[:19], fmt)
            except Exception:
                pass
    return None


def _wb_get_connection(db: Session, connection_id: uuid.UUID) -> models.MarketplaceConnection:
    conn = db.get(models.MarketplaceConnection, connection_id)
    if not conn:
        raise HTTPException(status_code=404, detail="Marketplace connection not found")
    if (conn.marketplace or "").lower() != "wb":
        raise HTTPException(status_code=400, detail="Connection marketplace must be 'wb'")
    if not conn.api_key:
        raise HTTPException(status_code=400, detail="WB token (api_key) is empty for this connection")
    return conn


def _wb_request(method: str, url: str, token: str, params: dict | None = None):
    with httpx.Client(timeout=60.0) as client:
        r = client.request(method, url, headers=_wb_headers(token), params=params)
        return r


@app.get("/integrations/wb/ping", response_model=schemas.WbPingOut)
def wb_ping(connection_id: uuid.UUID, db: Session = Depends(get_db)):
    conn = _wb_get_connection(db, connection_id)
    r = _wb_request("GET", f"{WB_STATS_BASE}/ping", conn.api_key)
    body = None
    try:
        body = r.text
    except Exception:
        body = None
    return schemas.WbPingOut(ok=r.status_code < 400, status_code=r.status_code, body=body)


@app.post("/integrations/wb/orders/fetch", response_model=schemas.WbFetchResult)
def wb_fetch_orders(params: schemas.WbFetchParams, db: Session = Depends(get_db)):
    conn = _wb_get_connection(db, params.connection_id)
    errors: list[str] = []

    # WB expects RFC3339-ish. Date only is allowed.
    qparams = {
        "dateFrom": params.date_from.isoformat(),
        "flag": 0,
    }
    r = _wb_request("GET", f"{WB_STATS_BASE}/api/v1/supplier/orders", conn.api_key, params=qparams)
    if r.status_code >= 400:
        raise HTTPException(status_code=400, detail=f"WB API error {r.status_code}: {r.text[:500]}")

    try:
        rows = r.json()
    except Exception:
        raise HTTPException(status_code=400, detail=f"WB API returned non-JSON: {r.text[:200]}")

    if not isinstance(rows, list):
        raise HTTPException(status_code=400, detail=f"WB API returned unexpected payload: {rows}")

    # Preload existing keys (srid + nmId + barcode) to split inserted/updated
    existing = set(
        db.execute(
            select(models.WbOrderLine.srid, models.WbOrderLine.nm_id, models.WbOrderLine.barcode)
            .where(models.WbOrderLine.connection_id == params.connection_id)
        ).all()
    )

    values = []
    inserted = 0
    updated = 0

    for it in rows:
        if not isinstance(it, dict):
            continue
        srid = str(it.get("srid") or "").strip()
        if not srid:
            continue
        nm_id = it.get("nmId")
        try:
            nm_id_int = int(nm_id) if nm_id is not None else None
        except Exception:
            nm_id_int = None
        barcode = it.get("barcode")
        barcode_str = str(barcode).strip() if barcode is not None else None

        key = (srid, nm_id_int, barcode_str)
        if key in existing:
            updated += 1
        else:
            inserted += 1
            existing.add(key)

        values.append(
            {
                "id": uuid.uuid4(),
                "connection_id": params.connection_id,
                "srid": srid,
                "nm_id": nm_id_int,
                "barcode": barcode_str,
                "supplier_article": (str(it.get("supplierArticle")).strip() if it.get("supplierArticle") is not None else None),
                "warehouse_name": (str(it.get("warehouseName")).strip() if it.get("warehouseName") is not None else None),
                "date": _wb_parse_dt(it.get("date")),
                "last_change_date": _wb_parse_dt(it.get("lastChangeDate")),
                "quantity": (int(it.get("quantity")) if it.get("quantity") is not None else None),
                "total_price": (float(it.get("totalPrice")) if it.get("totalPrice") is not None else None),
                "finished_price": (float(it.get("finishedPrice")) if it.get("finishedPrice") is not None else None),
                "price_with_disc": (float(it.get("priceWithDisc")) if it.get("priceWithDisc") is not None else None),
                "is_cancel": (bool(it.get("isCancel")) if it.get("isCancel") is not None else None),
                "cancel_date": _wb_parse_dt(it.get("cancelDate")),
                "raw_payload": it,
                "imported_at": datetime.utcnow(),
            }
        )

    if values:
        if engine.dialect.name == "postgresql":
            stmt = (
                pg_insert(models.WbOrderLine.__table__)
                .values(values)
                .on_conflict_do_update(
                    index_elements=["connection_id", "srid", "nm_id", "barcode"],
                    set_={
                        "supplier_article": text("EXCLUDED.supplier_article"),
                        "warehouse_name": text("EXCLUDED.warehouse_name"),
                        "date": text("EXCLUDED.date"),
                        "last_change_date": text("EXCLUDED.last_change_date"),
                        "quantity": text("EXCLUDED.quantity"),
                        "total_price": text("EXCLUDED.total_price"),
                        "finished_price": text("EXCLUDED.finished_price"),
                        "price_with_disc": text("EXCLUDED.price_with_disc"),
                        "is_cancel": text("EXCLUDED.is_cancel"),
                        "cancel_date": text("EXCLUDED.cancel_date"),
                        "raw_payload": text("EXCLUDED.raw_payload"),
                        "imported_at": text("EXCLUDED.imported_at"),
                    },
                )
            )
            db.execute(stmt)
        else:
            # SQLite fallback (slow but ok for MVP)
            for v in values:
                q = select(models.WbOrderLine).where(
                    and_(
                        models.WbOrderLine.connection_id == v["connection_id"],
                        models.WbOrderLine.srid == v["srid"],
                        models.WbOrderLine.nm_id == v["nm_id"],
                        models.WbOrderLine.barcode == v["barcode"],
                    )
                )
                ex = db.execute(q).scalars().first()
                if ex:
                    for k, val in v.items():
                        if k == "id":
                            continue
                        setattr(ex, k, val)
                else:
                    db.add(models.WbOrderLine(**v))
        db.commit()

    return schemas.WbFetchResult(fetched=len(rows), inserted=inserted, updated=updated, errors=errors)


@app.post("/integrations/wb/sales/fetch", response_model=schemas.WbFetchResult)
def wb_fetch_sales(params: schemas.WbFetchParams, db: Session = Depends(get_db)):
    conn = _wb_get_connection(db, params.connection_id)
    errors: list[str] = []

    qparams = {
        "dateFrom": params.date_from.isoformat(),
        "flag": 0,
    }
    r = _wb_request("GET", f"{WB_STATS_BASE}/api/v1/supplier/sales", conn.api_key, params=qparams)
    if r.status_code >= 400:
        raise HTTPException(status_code=400, detail=f"WB API error {r.status_code}: {r.text[:500]}")
    try:
        rows = r.json()
    except Exception:
        raise HTTPException(status_code=400, detail=f"WB API returned non-JSON: {r.text[:200]}")
    if not isinstance(rows, list):
        raise HTTPException(status_code=400, detail=f"WB API returned unexpected payload: {rows}")

    existing = set(
        db.execute(
            select(models.WbSaleLine.sale_id).where(models.WbSaleLine.connection_id == params.connection_id)
        ).scalars().all()
    )

    values = []
    inserted = 0
    updated = 0

    for it in rows:
        if not isinstance(it, dict):
            continue
        sale_id = str(it.get("saleID") or "").strip()
        if not sale_id:
            continue
        if sale_id in existing:
            updated += 1
        else:
            inserted += 1
            existing.add(sale_id)

        nm_id = it.get("nmId")
        try:
            nm_id_int = int(nm_id) if nm_id is not None else None
        except Exception:
            nm_id_int = None

        values.append(
            {
                "id": uuid.uuid4(),
                "connection_id": params.connection_id,
                "sale_id": sale_id,
                "srid": (str(it.get("srid")).strip() if it.get("srid") is not None else None),
                "nm_id": nm_id_int,
                "barcode": (str(it.get("barcode")).strip() if it.get("barcode") is not None else None),
                "supplier_article": (str(it.get("supplierArticle")).strip() if it.get("supplierArticle") is not None else None),
                "warehouse_name": (str(it.get("warehouseName")).strip() if it.get("warehouseName") is not None else None),
                "date": _wb_parse_dt(it.get("date")),
                "last_change_date": _wb_parse_dt(it.get("lastChangeDate")),
                "quantity": (int(it.get("quantity")) if it.get("quantity") is not None else None),
                "for_pay": (float(it.get("forPay")) if it.get("forPay") is not None else None),
                "finished_price": (float(it.get("finishedPrice")) if it.get("finishedPrice") is not None else None),
                "price_with_disc": (float(it.get("priceWithDisc")) if it.get("priceWithDisc") is not None else None),
                "raw_payload": it,
                "imported_at": datetime.utcnow(),
            }
        )

    if values:
        if engine.dialect.name == "postgresql":
            stmt = (
                pg_insert(models.WbSaleLine.__table__)
                .values(values)
                .on_conflict_do_update(
                    index_elements=["connection_id", "sale_id"],
                    set_={
                        "srid": text("EXCLUDED.srid"),
                        "nm_id": text("EXCLUDED.nm_id"),
                        "barcode": text("EXCLUDED.barcode"),
                        "supplier_article": text("EXCLUDED.supplier_article"),
                        "warehouse_name": text("EXCLUDED.warehouse_name"),
                        "date": text("EXCLUDED.date"),
                        "last_change_date": text("EXCLUDED.last_change_date"),
                        "quantity": text("EXCLUDED.quantity"),
                        "for_pay": text("EXCLUDED.for_pay"),
                        "finished_price": text("EXCLUDED.finished_price"),
                        "price_with_disc": text("EXCLUDED.price_with_disc"),
                        "raw_payload": text("EXCLUDED.raw_payload"),
                        "imported_at": text("EXCLUDED.imported_at"),
                    },
                )
            )
            db.execute(stmt)
        else:
            for v in values:
                ex = db.execute(
                    select(models.WbSaleLine).where(
                        and_(models.WbSaleLine.connection_id == v["connection_id"], models.WbSaleLine.sale_id == v["sale_id"])
                    )
                ).scalars().first()
                if ex:
                    for k, val in v.items():
                        if k == "id":
                            continue
                        setattr(ex, k, val)
                else:
                    db.add(models.WbSaleLine(**v))
        db.commit()

    return schemas.WbFetchResult(fetched=len(rows), inserted=inserted, updated=updated, errors=errors)


@app.get("/integrations/wb/orders", response_model=list[schemas.WbOrderLineOut])
def wb_list_orders(
    connection_id: uuid.UUID,
    date_from: date | None = None,
    date_to: date | None = None,
    limit: int = 500,
    db: Session = Depends(get_db),
):
    _ = _wb_get_connection(db, connection_id)
    q = select(models.WbOrderLine).where(models.WbOrderLine.connection_id == connection_id)
    if date_from:
        q = q.where(models.WbOrderLine.date >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.where(models.WbOrderLine.date <= datetime.combine(date_to, datetime.max.time()))
    q = q.order_by(models.WbOrderLine.date.desc().nullslast()).limit(max(1, min(2000, limit)))
    return list(db.execute(q).scalars().all())


@app.get("/integrations/wb/sales", response_model=list[schemas.WbSaleLineOut])
def wb_list_sales(
    connection_id: uuid.UUID,
    date_from: date | None = None,
    date_to: date | None = None,
    limit: int = 500,
    db: Session = Depends(get_db),
):
    _ = _wb_get_connection(db, connection_id)
    q = select(models.WbSaleLine).where(models.WbSaleLine.connection_id == connection_id)
    if date_from:
        q = q.where(models.WbSaleLine.date >= datetime.combine(date_from, datetime.min.time()))
    if date_to:
        q = q.where(models.WbSaleLine.date <= datetime.combine(date_to, datetime.max.time()))
    q = q.order_by(models.WbSaleLine.date.desc().nullslast()).limit(max(1, min(2000, limit)))
    return list(db.execute(q).scalars().all())


# ---------------------------
# FBS Builds (internal batches)
# ---------------------------


_FBS_BUILD_STATUSES = {"draft", "picking", "packed", "shipped", "closed", "cancelled"}


def get_marketplace_connection(db: Session, connection_id: uuid.UUID) -> models.MarketplaceConnection:
    conn = db.get(models.MarketplaceConnection, connection_id)
    if not conn:
        raise HTTPException(status_code=404, detail="connection not found")
    return conn


def _get_fbs_order_snapshot(db: Session, marketplace: str, connection_id: uuid.UUID, external_order_id: str) -> dict:
    """Return {status, order_payload, items_payload} for an order id."""

    marketplace = (marketplace or "").lower().strip()

    if marketplace == "ozon":
        posting = db.execute(
            select(models.OzonPosting)
            .options(selectinload(models.OzonPosting.items))
            .where(
                and_(
                    models.OzonPosting.connection_id == connection_id,
                    models.OzonPosting.posting_number == external_order_id,
                )
            )
        ).scalars().first()
        if not posting:
            raise HTTPException(status_code=404, detail=f"Ozon posting not found: {external_order_id}")
        items = []
        for it in (posting.items or []):
            items.append(
                {
                    "sku": it.sku,
                    "offer_id": it.offer_id,
                    "name": it.name,
                    "qty": int(it.quantity or 0),
                    "price": float(it.price or 0),
                }
            )
        return {
            "status": posting.status,
            "order_payload": posting.raw_payload or {},
            "items_payload": items,
        }

    if marketplace == "ymarket":
        try:
            oid = int(str(external_order_id))
        except Exception:
            raise HTTPException(status_code=400, detail="YMarket order_id должен быть числом")
        order = db.execute(
            select(models.YMarketOrder)
            .options(selectinload(models.YMarketOrder.items))
            .where(and_(models.YMarketOrder.connection_id == connection_id, models.YMarketOrder.order_id == oid))
        ).scalars().first()
        if not order:
            raise HTTPException(status_code=404, detail=f"YMarket order not found: {external_order_id}")
        items = []
        for it in (order.items or []):
            sku = it.shop_sku or it.offer_id or (str(it.market_sku) if it.market_sku else None)
            items.append(
                {
                    "sku": sku,
                    "offer_id": it.offer_id,
                    "name": it.name,
                    "qty": int(it.quantity or 0),
                    "price": float(it.price or 0),
                }
            )
        return {
            "status": order.status,
            "order_payload": order.raw_payload or {},
            "items_payload": items,
        }

    if marketplace == "wb":
        lines = list(
            db.execute(
                select(models.WbOrderLine).where(
                    and_(models.WbOrderLine.connection_id == connection_id, models.WbOrderLine.srid == external_order_id)
                )
            ).scalars().all()
        )
        if not lines:
            raise HTTPException(status_code=404, detail=f"WB order(srid) not found: {external_order_id}")
        by_sku: dict[str, dict] = {}
        for ln in lines:
            sku = ln.supplier_article or ln.barcode or (str(ln.nm_id) if ln.nm_id is not None else None) or "—"
            rec = by_sku.get(sku)
            if not rec:
                rec = {"sku": sku, "offer_id": None, "name": sku, "qty": 0, "price": 0.0}
                by_sku[sku] = rec
            rec["qty"] += int(ln.quantity or 0)
            rec["price"] = float(ln.price_with_disc or ln.total_price or 0)
        status = "cancelled" if any(bool(x.is_cancel) for x in lines) else "new"
        return {
            "status": status,
            "order_payload": {"srid": external_order_id},
            "items_payload": list(by_sku.values()),
        }

    raise HTTPException(status_code=400, detail="Unsupported marketplace")


def _build_counts(orders: list[models.FbsBuildOrder]) -> tuple[int, int, int]:
    orders_count = len(orders)
    items_count = 0
    qty_total = 0
    for o in orders:
        items = (o.items_payload or []) if isinstance(o.items_payload, list) else []
        items_count += len(items)
        for it in items:
            try:
                qty_total += int((it or {}).get("qty") or 0)
            except Exception:
                pass
    return orders_count, items_count, qty_total


@app.get("/integrations/fbs/builds", response_model=list[schemas.FbsBuildOut])
def fbs_list_builds(
    marketplace: str,
    connection_id: uuid.UUID,
    limit: int = 200,
    db: Session = Depends(get_db),
):
    mp = (marketplace or "").lower().strip()
    _ = get_marketplace_connection(db, connection_id)
    q = (
        select(models.FbsBuild)
        .options(selectinload(models.FbsBuild.orders))
        .where(and_(models.FbsBuild.marketplace == mp, models.FbsBuild.connection_id == connection_id))
        .order_by(models.FbsBuild.created_at.desc())
        .limit(max(1, min(500, limit)))
    )
    builds = list(db.execute(q).scalars().all())

    out: list[schemas.FbsBuildOut] = []
    for b in builds:
        oc, ic, qt = _build_counts(list(b.orders or []))
        out.append(
            schemas.FbsBuildOut(
                id=b.id,
                marketplace=b.marketplace,
                connection_id=b.connection_id,
                title=b.title,
                status=b.status,
                note=b.note,
                created_at=b.created_at,
                updated_at=b.updated_at,
                orders_count=oc,
                items_count=ic,
                qty_total=qt,
            )
        )
    return out


@app.post("/integrations/fbs/builds", response_model=schemas.FbsBuildDetailOut)
def fbs_create_build(payload: schemas.FbsBuildCreateParams, db: Session = Depends(get_db)):
    mp = (payload.marketplace or "").lower().strip()
    if mp not in {"ozon", "ymarket", "wb"}:
        raise HTTPException(status_code=400, detail="marketplace must be ozon|ymarket|wb")
    conn = get_marketplace_connection(db, payload.connection_id)
    if (conn.marketplace or "").lower().strip() != mp:
        raise HTTPException(status_code=400, detail="connection.marketplace != payload.marketplace")

    title = (payload.title or "").strip() or f"Сборка {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}"
    b = models.FbsBuild(
        marketplace=mp,
        connection_id=payload.connection_id,
        title=title,
        status="draft",
        note=(payload.note or None),
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
    )
    db.add(b)
    db.flush()

    uniq_ids = []
    seen = set()
    for oid in (payload.order_ids or []):
        s = str(oid).strip()
        if not s or s in seen:
            continue
        seen.add(s)
        uniq_ids.append(s)

    for oid in uniq_ids:
        snap = _get_fbs_order_snapshot(db, mp, payload.connection_id, oid)
        o = models.FbsBuildOrder(
            build_id=b.id,
            marketplace=mp,
            connection_id=payload.connection_id,
            external_order_id=oid,
            status=snap.get("status"),
            order_payload=snap.get("order_payload"),
            items_payload=snap.get("items_payload"),
            created_at=datetime.utcnow(),
        )
        db.add(o)

    b.updated_at = datetime.utcnow()
    db.commit()
    return fbs_get_build(b.id, db)


@app.get("/integrations/fbs/builds/{build_id}", response_model=schemas.FbsBuildDetailOut)
def fbs_get_build(build_id: uuid.UUID, db: Session = Depends(get_db)):
    b = db.execute(
        select(models.FbsBuild)
        .options(selectinload(models.FbsBuild.orders))
        .where(models.FbsBuild.id == build_id)
    ).scalars().first()
    if not b:
        raise HTTPException(status_code=404, detail="build not found")

    orders = list(b.orders or [])
    oc, ic, qt = _build_counts(orders)

    # Aggregate items by (sku, offer_id, name)
    agg: dict[tuple, dict] = {}
    for o in orders:
        items = (o.items_payload or []) if isinstance(o.items_payload, list) else []
        for it in items:
            sku = (it or {}).get("sku")
            offer_id = (it or {}).get("offer_id")
            name = (it or {}).get("name")
            key = (sku, offer_id, name)
            rec = agg.get(key)
            if not rec:
                rec = {"sku": sku, "offer_id": offer_id, "name": name, "qty_total": 0, "orders": set()}
                agg[key] = rec
            try:
                rec["qty_total"] += int((it or {}).get("qty") or 0)
            except Exception:
                pass
            rec["orders"].add(o.external_order_id)

    items_out = []
    for rec in agg.values():
        items_out.append(
            schemas.FbsBuildItemAggOut(
                sku=rec.get("sku"),
                offer_id=rec.get("offer_id"),
                name=rec.get("name"),
                qty_total=int(rec.get("qty_total") or 0),
                orders_count=len(rec.get("orders") or []),
            )
        )
    items_out.sort(key=lambda x: (-(x.qty_total or 0), str(x.sku or "")))

    orders_out = []
    for o in orders:
        items = (o.items_payload or []) if isinstance(o.items_payload, list) else []
        qty = 0
        for it in items:
            try:
                qty += int((it or {}).get("qty") or 0)
            except Exception:
                pass
        orders_out.append(
            schemas.FbsBuildOrderOut(
                id=o.id,
                external_order_id=o.external_order_id,
                status=o.status,
                items_count=len(items),
                qty_total=qty,
                items=items,
            )
        )

    return schemas.FbsBuildDetailOut(
        id=b.id,
        marketplace=b.marketplace,
        connection_id=b.connection_id,
        title=b.title,
        status=b.status,
        note=b.note,
        created_at=b.created_at,
        updated_at=b.updated_at,
        orders_count=oc,
        items_count=ic,
        qty_total=qt,
        orders=orders_out,
        items=items_out,
    )


@app.patch("/integrations/fbs/builds/{build_id}", response_model=schemas.FbsBuildDetailOut)
def fbs_patch_build(build_id: uuid.UUID, payload: schemas.FbsBuildPatchParams, db: Session = Depends(get_db)):
    b = db.execute(select(models.FbsBuild).where(models.FbsBuild.id == build_id)).scalars().first()
    if not b:
        raise HTTPException(status_code=404, detail="build not found")
    if payload.title is not None:
        b.title = payload.title.strip() or b.title
    if payload.note is not None:
        b.note = payload.note
    if payload.status is not None:
        st = payload.status.strip().lower()
        if st not in _FBS_BUILD_STATUSES:
            raise HTTPException(status_code=400, detail="invalid status")
        b.status = st
    b.updated_at = datetime.utcnow()
    db.commit()
    return fbs_get_build(build_id, db)


@app.post("/integrations/fbs/builds/{build_id}/add_orders", response_model=schemas.FbsBuildDetailOut)
def fbs_add_orders(build_id: uuid.UUID, payload: schemas.FbsBuildAddOrdersParams, db: Session = Depends(get_db)):
    b = db.execute(
        select(models.FbsBuild).options(selectinload(models.FbsBuild.orders)).where(models.FbsBuild.id == build_id)
    ).scalars().first()
    if not b:
        raise HTTPException(status_code=404, detail="build not found")

    existing = {o.external_order_id for o in (b.orders or [])}
    for oid in (payload.order_ids or []):
        s = str(oid).strip()
        if not s or s in existing:
            continue
        snap = _get_fbs_order_snapshot(db, b.marketplace, b.connection_id, s)
        o = models.FbsBuildOrder(
            build_id=b.id,
            marketplace=b.marketplace,
            connection_id=b.connection_id,
            external_order_id=s,
            status=snap.get("status"),
            order_payload=snap.get("order_payload"),
            items_payload=snap.get("items_payload"),
            created_at=datetime.utcnow(),
        )
        db.add(o)

    b.updated_at = datetime.utcnow()
    db.commit()
    return fbs_get_build(build_id, db)
